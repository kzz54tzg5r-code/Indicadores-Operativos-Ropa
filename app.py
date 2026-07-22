
# -*- coding: utf-8 -*-
import base64
import hashlib
import json
import re
import sqlite3
import unicodedata
from datetime import datetime, date
from pathlib import Path
from zoneinfo import ZoneInfo
from urllib.parse import quote

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing, String, PolyLine, Circle, Rect, Line
from reportlab.graphics.charts.barcharts import VerticalBarChart
import streamlit as st
from openpyxl import load_workbook

try:
    from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
    AGGRID_OK = True
except Exception:
    AGGRID_OK = False


# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================
st.set_page_config(
    page_title="Indicadores Cambios y Muertos",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

DATA_DIR = Path("data")
UPLOAD_DIR = DATA_DIR / "uploads"
CACHE_DIR = DATA_DIR / "cache"
CONFIG_DIR = DATA_DIR / "config"
ASSETS_DIR = Path("assets")
ACTIVE_FILE = UPLOAD_DIR / "base_activa.xlsx"
META_FILE = CONFIG_DIR / "metadata.json"
DB_FILE = CONFIG_DIR / "usuarios.db"

for p in [DATA_DIR, UPLOAD_DIR, CACHE_DIR, CONFIG_DIR, ASSETS_DIR]:
    p.mkdir(parents=True, exist_ok=True)

MX_TZ = ZoneInfo("America/Mexico_City")
APP_CACHE_VERSION = "v12.0"
AZUL = "#10245F"
ROSA = "#EC007C"
LAVANDA = "#F3F6FB"

PROJECT_STORES = [
    "Arco Norte", "Ecatepec", "Miravalle", "Puebla Sur", "Vallejo",
]


# ============================================================
# UTILIDADES
# ============================================================
def norm_text(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"\s+", " ", s)
    return s.upper().strip()


STORE_MAP = {
    "ARCO NORTE": "Arco Norte",
    "ECATEPEC": "Ecatepec",
    "MIRAVALLE": "Miravalle",
    "PUEBLA SUR": "Puebla Sur",
    "VALLEJO": "Vallejo",
    "PUEBLA": "Puebla",
    "IZTAPALAPA": "Iztapalapa",
    "TOLUCA": "Toluca",
    "CENTRO": "Centro",
    "QUERETARO": "Querétaro",
    "QUERÉTARO": "Querétaro",
    "LEON": "León",
    "LEÓN": "León",
    "NAUCALPAN": "Naucalpan",
    "OLIVAR": "Olivar",
    "AGUASCALIENTES": "Aguascalientes",
    "VERACRUZ": "Veracruz",
    "IXTAPALUCA": "Ixtapaluca",
    "VALLEJO ": "Vallejo",
}


def canon_store(x):
    if pd.isna(x):
        return ""
    raw = str(x).strip()
    if not raw:
        return ""

    s = norm_text(raw)
    s_clean = re.sub(r"[^A-Z0-9]+", " ", s).strip()

    if "MIRAVALLE" in s_clean:
        return "Miravalle"
    if "ATEMAJAC" in s_clean:
        return "Atemajac"
    if s_clean in ["GUADALAJARA", "GDL", "GUADALAJARA JALISCO"]:
        return "Atemajac"

    if "ARCO" in s_clean and "NORTE" in s_clean:
        return "Arco Norte"
    if "PUEBLA" in s_clean and "SUR" in s_clean:
        return "Puebla Sur"
    if s_clean in ["PUEBLA CENTRO", "PUEBLA CENTRO ROPA"] or s_clean == "PUEBLA":
        return "Puebla"
    if "ECATEPEC" in s_clean:
        return "Ecatepec"
    if "VALLEJO" in s_clean:
        return "Vallejo"
    if "IZTAPALAPA" in s_clean:
        return "Iztapalapa"
    if "IXTAPALUCA" in s_clean:
        return "Ixtapaluca"
    if "NAUCALPAN" in s_clean:
        return "Naucalpan"
    if "TOLUCA" in s_clean:
        return "Toluca"
    if "QUERETARO" in s_clean or "QUERÉTARO" in raw.upper():
        return "Querétaro"
    if "LEON" in s_clean or "LEÓN" in raw.upper():
        return "León"
    if "VERACRUZ" in s_clean:
        return "Veracruz"
    if "AGUASCALIENTES" in s_clean:
        return "Aguascalientes"
    if "OLIVAR" in s_clean:
        return "Olivar"
    if "SAN LUIS" in s_clean:
        return "San Luis"
    if s_clean == "CENTRO" or "CENTRO HISTORICO" in s_clean or "CENTRO HISTÓRICO" in raw.upper():
        return "Centro"

    try:
        for k, v in STORE_MAP.items():
            if norm_text(k) == s or norm_text(k) == s_clean:
                return v
    except Exception:
        pass

    invalid = {
        "TIENDA", "TIENDAS", "DIA", "DÍA", "FECHA", "VENTAS NETA PZS", "VENTAS NETAS",
        "DEV PZS", "VENTA NETA EN", "VENTA NETA", "CATEGORIA", "SUB CATEGORIA",
        "SUB CATEGORÍA", "FAMILIA RLN", "GRUPO RLN", "PRECIO MENUDEO"
    }
    if s_clean in invalid or s in invalid:
        return ""

    return raw.title()



def safe_num(x) -> float:
    if pd.isna(x):
        return 0.0
    s = str(x).strip().replace("$", "").replace(",", "").replace(" ", "")
    if s in ["", "-", "nan", "None"]:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0



def excel_col_name(n):
    """Convierte índice 0-based a letra de Excel."""
    n = int(n) + 1
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def parse_date(x):
    """Convierte fechas de Excel/UI a Timestamp normalizado sin invertir mes y día.

    Casos soportados:
    - 2026-07-09
    - 2026/07/09
    - 2026-07-09 17:00:14
    - 2026/07/09 17:00:14
    - 09/07/2026
    - Fechas reales de Excel
    - Números seriales de Excel
    """
    if x is None:
        return pd.NaT

    try:
        if pd.isna(x):
            return pd.NaT
    except Exception:
        pass

    # Objetos de fecha reales.
    if isinstance(x, (pd.Timestamp, datetime, date)):
        try:
            return pd.Timestamp(x).normalize()
        except Exception:
            return pd.NaT

    # Seriales de Excel o timestamps numéricos.
    if isinstance(x, (int, float, np.integer, np.floating)):
        val = float(x)
        if not np.isfinite(val):
            return pd.NaT

        if 20000 <= val <= 60000:
            try:
                return (
                    pd.Timestamp("1899-12-30")
                    + pd.to_timedelta(val, unit="D")
                ).normalize()
            except Exception:
                return pd.NaT

        for unit, minimum in [("ns", 10**14), ("ms", 10**11), ("s", 10**9)]:
            if val > minimum:
                try:
                    parsed = pd.to_datetime(int(val), unit=unit, errors="coerce")
                    return parsed.normalize() if pd.notna(parsed) else pd.NaT
                except Exception:
                    pass

    s = str(x).strip()
    if not s or s in {"-", "nan", "NaT", "None"}:
        return pd.NaT

    # Año primero, con o sin hora. Este es el formato de Resultados productividad 2.
    # Ejemplo: 2026-07-09 17:00:14.
    if re.match(
        r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}(?:[ T]\d{1,2}:\d{2}(?::\d{2}(?:\.\d+)?)?)?$",
        s,
    ):
        parsed = pd.to_datetime(
            s,
            errors="coerce",
            yearfirst=True,
            dayfirst=False,
        )
        return parsed.normalize() if pd.notna(parsed) else pd.NaT

    # Día primero, con o sin hora.
    # Ejemplo: 09/07/2026 17:00:14.
    if re.match(
        r"^\d{1,2}[-/]\d{1,2}[-/]\d{4}(?:[ T]\d{1,2}:\d{2}(?::\d{2}(?:\.\d+)?)?)?$",
        s,
    ):
        parsed = pd.to_datetime(
            s,
            errors="coerce",
            dayfirst=True,
            yearfirst=False,
        )
        return parsed.normalize() if pd.notna(parsed) else pd.NaT

    # Posible serial guardado como texto.
    compact = s.replace("$", "").replace(",", "").replace(" ", "")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", compact):
        try:
            val = float(compact)
            if 20000 <= val <= 60000:
                return (
                    pd.Timestamp("1899-12-30")
                    + pd.to_timedelta(val, unit="D")
                ).normalize()
        except Exception:
            pass

    # Último intento controlado. Se prueba primero año-mes-día y después día-mes-año.
    parsed = pd.to_datetime(
        s,
        errors="coerce",
        yearfirst=True,
        dayfirst=False,
    )
    if pd.isna(parsed):
        parsed = pd.to_datetime(
            s,
            errors="coerce",
            dayfirst=True,
            yearfirst=False,
        )

    return parsed.normalize() if pd.notna(parsed) else pd.NaT


def fmt_num(x):
    return f"{safe_num(x):,.0f}"


def fmt_money(x):
    return f"${safe_num(x):,.0f}"


def fmt_pct(x):
    return f"{safe_num(x):.1f}%"


def hash_password(pwd):
    return hashlib.sha256(str(pwd).encode("utf-8")).hexdigest()


def logo_html():
    logo = ASSETS_DIR / "price_shoes_logo.png"
    if logo.exists():
        data = base64.b64encode(logo.read_bytes()).decode("utf-8")
        return f'<img src="data:image/png;base64,{data}" class="ps-logo-img">'
    return '<div class="ps-logo-text">Price<br>Shoes</div>'


# ============================================================
# USUARIOS
# ============================================================
def init_db():
    con = sqlite3.connect(DB_FILE)
    cur = con.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS usuarios (
            nomina TEXT PRIMARY KEY,
            nombre TEXT NOT NULL,
            permiso TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            activo INTEGER NOT NULL DEFAULT 1,
            creado TEXT
        )
        """
    )
    cur.execute("SELECT COUNT(*) FROM usuarios")
    if cur.fetchone()[0] == 0:
        cur.execute(
            "INSERT INTO usuarios VALUES (?,?,?,?,?,?)",
            ("admin", "Administrador", "Administrador", hash_password("admin123"), 1, datetime.now(MX_TZ).isoformat()),
        )
    con.commit()
    con.close()


def get_user(nomina, password):
    con = sqlite3.connect(DB_FILE)
    cur = con.cursor()
    cur.execute("SELECT nomina,nombre,permiso FROM usuarios WHERE nomina=? AND password_hash=? AND activo=1", (nomina, hash_password(password)))
    row = cur.fetchone()
    con.close()
    if not row:
        return None
    return {"nomina": row[0], "nombre": row[1], "permiso": row[2]}


def upsert_user(nomina, nombre, permiso, password):
    con = sqlite3.connect(DB_FILE)
    cur = con.cursor()
    cur.execute(
        """
        INSERT INTO usuarios(nomina,nombre,permiso,password_hash,activo,creado)
        VALUES (?,?,?,?,1,?)
        ON CONFLICT(nomina) DO UPDATE SET
            nombre=excluded.nombre,
            permiso=excluded.permiso,
            password_hash=excluded.password_hash,
            activo=1
        """,
        (str(nomina), nombre, permiso, hash_password(password), datetime.now(MX_TZ).isoformat()),
    )
    con.commit()
    con.close()


def delete_user(nomina):
    if nomina == "admin":
        return
    con = sqlite3.connect(DB_FILE)
    con.execute("DELETE FROM usuarios WHERE nomina=?", (nomina,))
    con.commit()
    con.close()


def list_users():
    con = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("SELECT nomina AS Nómina, nombre AS Nombre, permiso AS Permiso, activo AS Activo FROM usuarios ORDER BY nombre", con)
    con.close()
    return df


init_db()


# ============================================================
# ESTILOS
# ============================================================
def apply_styles():
    st.markdown(
        f"""
<style>
:root {{
    --azul:{AZUL};
    --rosa:{ROSA};
}}
html, body, [data-testid="stAppViewContainer"] {{
    background:#F3F6FB;
}}
.block-container {{
    padding-top:0.8rem!important;
    padding-left:1.6rem!important;
    padding-right:1.6rem!important;
    max-width:100%!important;
}}
.ps-top-line {{
    height:6px;
    background:{ROSA};
    margin:0 -1.6rem 18px -1.6rem;
}}
.ps-header {{
    width:100%;
    background:#FFF;
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:22px;
    padding:14px 24px 18px 24px;
    box-sizing:border-box;
}}
.ps-header-left {{
    display:flex;
    align-items:center;
    gap:24px;
    min-width:0;
}}
.ps-logo-wrap {{
    width:126px;
    height:82px;
    display:flex;
    align-items:center;
    justify-content:center;
}}
.ps-logo-img {{
    max-width:120px!important;
    max-height:78px!important;
    object-fit:contain!important;
}}
.ps-logo-text {{
    color:{AZUL};
    font-weight:900;
    font-size:26px;
    line-height:1;
}}
.ps-header-sep {{
    width:5px;
    height:86px;
    background:{ROSA};
    border-radius:3px;
}}
.ps-title {{
    color:#1D1259;
    font-weight:900;
    font-size:33px;
    line-height:1.08;
}}
.ps-subtitle {{
    color:#5B6476;
    font-weight:800;
    font-size:15px;
    margin-top:7px;
}}
.ps-header-right {{
    display:flex;
    gap:14px;
    align-items:center;
}}
.ps-meta {{
    min-width:185px;
    background:#F8FAFC;
    border:1px solid #DDE4F0;
    border-radius:0 0 14px 14px;
    padding:12px 16px;
}}
.ps-meta-label {{
    color:#6B7280;
    letter-spacing:5px;
    font-size:12px;
    font-weight:900;
}}
.ps-meta-value {{
    color:#1D1259;
    font-size:18px;
    font-weight:900;
    margin-top:6px;
}}
.ps-tabbar {{
    background:{AZUL};
    border-top:4px solid {ROSA};
    margin:0 -1.6rem 22px -1.6rem;
    padding:0 70px;
    overflow-x:auto;
    white-space:nowrap;
}}
.ps-tabbar [role="radiogroup"] {{
    display:flex!important;
    flex-wrap:nowrap!important;
    gap:0!important;
    min-height:58px!important;
}}
.ps-tabbar label {{
    background:{AZUL}!important;
    color:#C7D2FE!important;
    min-height:58px!important;
    padding:0 18px!important;
    display:flex!important;
    align-items:center!important;
    border-radius:0!important;
    font-weight:900!important;
    white-space:nowrap!important;
}}
.ps-tabbar label:hover {{
    background:#142E73!important;
    color:#FFF!important;
}}
.ps-tabbar label:has(input:checked) {{
    background:#142E73!important;
    color:#FFF!important;
    border-bottom:4px solid {ROSA}!important;
}}
.ps-tabbar label * {{
    color:inherit!important;
    font-weight:900!important;
}}
.ps-kpi-grid {{
    display:grid;
    grid-template-columns:repeat(auto-fit,minmax(250px,1fr));
    gap:18px;
    margin:18px 0 22px 0;
}}
.ps-kpi-card {{
    background:#FFF;
    border:1px solid #E1E7F0;
    border-radius:14px;
    padding:22px 18px;
    min-height:145px;
    display:flex;
    align-items:center;
    gap:18px;
    box-shadow:0 8px 20px rgba(16,36,95,.06);
    overflow:hidden;
}}
.ps-kpi-icon {{
    width:76px;
    height:76px;
    min-width:76px;
    border-radius:50%;
    color:#FFF;
    font-size:34px;
    display:flex;
    align-items:center;
    justify-content:center;
    font-weight:900;
}}
.ps-kpi-title {{
    color:#17132D;
    font-size:15px;
    font-weight:900;
    line-height:1.2;
}}
.ps-kpi-value {{
    color:{ROSA};
    font-size:30px;
    font-weight:900;
    line-height:1.1;
    margin:8px 0;
}}
.ps-kpi-sub {{
    color:#17132D;
    font-size:13px;
    line-height:1.35;
}}
.panel-title {{
    background:#FFF;
    border:1px solid #E1E7F0;
    border-radius:12px;
    padding:18px 22px;
    margin:18px 0 12px 0;
    font-size:20px;
    font-weight:900;
    color:#17132D;
}}
.ag-header,.ag-header-cell {{
    background:{AZUL}!important;
}}
.ag-header-cell-text,.ag-header-cell-label,.ag-icon {{
    color:#FFF!important;
    fill:#FFF!important;
    font-weight:900!important;
}}
.ag-root-wrapper {{
    border-radius:10px!important;
    border:1px solid #E1E7F0!important;
    overflow:hidden!important;
}}
.ag-cell {{
    font-size:12px!important;
}}
.stButton > button {{
    border-radius:8px!important;
}}
.footer {{
    color:#7A8190;
    font-size:13px;
    margin:36px 0 10px 0;
    border-top:1px solid #DDE4F0;
    padding-top:18px;
}}
@media(max-width:1200px) {{
    .ps-header{{flex-direction:column;align-items:flex-start;}}
    .ps-header-right{{flex-wrap:wrap;}}
    .ps-tabbar{{padding:0 20px;}}
}}

.week-card-grid{{
    display:grid;
    grid-template-columns:repeat(4,minmax(230px,1fr));
    gap:22px;
    margin:16px 0 28px 0;
}}
.week-card{{
    background:#F8F9FC;
    border:1px solid #D9DEE8;
    border-radius:8px;
    overflow:hidden;
    box-shadow:0 5px 14px rgba(16,36,95,.06);
}}
.week-card-head{{
    background:#3E4095;
    color:white;
    text-align:center;
    font-size:20px;
    font-weight:900;
    padding:15px 10px;
}}
.week-row{{
    display:grid;
    grid-template-columns:1fr auto 62px;
    align-items:center;
    gap:12px;
    padding:14px 16px;
    border-bottom:1px solid #E5E7EB;
}}
.week-row span{{
    color:#666;
    font-weight:900;
    font-size:13px;
}}
.week-row b{{
    color:#3E4095;
    font-size:20px;
    font-weight:900;
}}
.week-row em{{
    font-style:normal;
    font-weight:900;
    font-size:12px;
    text-align:right;
}}
@media(max-width:1200px){{
    .week-card-grid{{grid-template-columns:repeat(2,minmax(230px,1fr));}}
}}
@media(max-width:700px){{
    .week-card-grid{{grid-template-columns:1fr;}}
}}


/* Navegación corporativa de extremo a extremo */
.ps-tabbar {{
    position: relative !important;
    left: 50% !important;
    right: 50% !important;
    margin-left: -50vw !important;
    margin-right: -50vw !important;
    width: 100vw !important;
    max-width: 100vw !important;
    box-sizing: border-box !important;
    background: var(--azul) !important;
    border-top: 5px solid var(--rosa) !important;
    padding: 0 24px !important;
    overflow-x: auto !important;
}}
.ps-tabbar [role="radiogroup"] {{
    width: max-content !important;
    min-width: 100% !important;
    justify-content: flex-start !important;
    background: var(--azul) !important;
}}
.ps-tabbar label {{
    background: var(--azul) !important;
    color: rgba(255,255,255,.72) !important;
    border: 0 !important;
    border-radius: 0 !important;
    min-height: 58px !important;
    padding: 0 22px !important;
}}
.ps-tabbar label p,
.ps-tabbar label span {{
    color: rgba(255,255,255,.72) !important;
    font-weight: 800 !important;
}}
.ps-tabbar label:has(input:checked) {{
    background: #142E73 !important;
    box-shadow: inset 0 -5px 0 var(--rosa) !important;
}}
.ps-tabbar label:has(input:checked) p,
.ps-tabbar label:has(input:checked) span {{
    color: #FFFFFF !important;
    font-weight: 900 !important;
}}
.ps-tabbar input[type="radio"] {{
    accent-color: #FFFFFF !important;
}}


/* =========================================================
   VISTA RESPONSIVA: COMPUTADORA Y MÓVIL
   ========================================================= */

/* Computadora y tablet horizontal */
@media (min-width: 769px) {{
    .ps-kpi-grid {{
        grid-template-columns: repeat(5, minmax(0, 1fr)) !important;
    }}

    .ps-kpi-card {{
        min-width: 0 !important;
    }}
}}

/* Móvil y tablet vertical */
@media (max-width: 768px) {{
    html, body, [data-testid="stAppViewContainer"] {{
        overflow-x: hidden !important;
    }}

    .block-container {{
        padding-top: .35rem !important;
        padding-left: .65rem !important;
        padding-right: .65rem !important;
        padding-bottom: 1rem !important;
    }}

    .ps-top-line {{
        height: 4px !important;
        margin: 0 -.65rem 8px -.65rem !important;
    }}

    .ps-header {{
        padding: 8px 8px 10px 8px !important;
        gap: 8px !important;
        flex-direction: column !important;
        align-items: stretch !important;
    }}

    .ps-header-left {{
        gap: 9px !important;
        width: 100% !important;
    }}

    .ps-logo-wrap {{
        width: 58px !important;
        height: 50px !important;
        min-width: 58px !important;
    }}

    .ps-logo-img {{
        max-width: 56px !important;
        max-height: 46px !important;
    }}

    .ps-logo-text {{
        font-size: 15px !important;
    }}

    .ps-header-sep {{
        width: 3px !important;
        height: 48px !important;
        min-width: 3px !important;
    }}

    .ps-title {{
        font-size: 19px !important;
        line-height: 1.05 !important;
    }}

    .ps-subtitle {{
        font-size: 10px !important;
        line-height: 1.15 !important;
        margin-top: 3px !important;
    }}

    .ps-header-right {{
        display: grid !important;
        grid-template-columns: 1fr 1fr !important;
        gap: 7px !important;
        width: 100% !important;
    }}

    .ps-meta {{
        min-width: 0 !important;
        border-radius: 8px !important;
        padding: 7px 9px !important;
    }}

    .ps-meta-label {{
        letter-spacing: 2px !important;
        font-size: 8px !important;
    }}

    .ps-meta-value {{
        font-size: 12px !important;
        margin-top: 3px !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }}

    .ps-tabbar {{
        margin: 0 -.65rem 12px -.65rem !important;
        padding: 0 6px !important;
        overflow-x: auto !important;
        -webkit-overflow-scrolling: touch !important;
        scrollbar-width: thin !important;
    }}

    .ps-tabbar [role="radiogroup"] {{
        min-height: 44px !important;
        width: max-content !important;
    }}

    .ps-tabbar label {{
        min-height: 44px !important;
        padding: 0 11px !important;
        font-size: 10px !important;
    }}

    h1 {{
        font-size: 1.55rem !important;
    }}

    h2 {{
        font-size: 1.32rem !important;
    }}

    h3 {{
        font-size: 1.1rem !important;
    }}

    .ps-kpi-grid {{
        grid-template-columns: repeat(2, minmax(0, 1fr)) !important;
        gap: 9px !important;
        margin: 10px 0 14px 0 !important;
    }}

    .ps-kpi-card {{
        min-height: 112px !important;
        height: auto !important;
        padding: 11px 9px !important;
        border-radius: 12px !important;
        gap: 8px !important;
        align-items: center !important;
        box-shadow: 0 4px 12px rgba(16,36,95,.06) !important;
    }}

    .ps-kpi-card:nth-child(5) {{
        grid-column: 1 / -1 !important;
        min-height: 96px !important;
    }}

    .ps-kpi-icon {{
        width: 47px !important;
        height: 47px !important;
        min-width: 47px !important;
        font-size: 23px !important;
    }}

    .ps-kpi-title {{
        font-size: 11px !important;
        line-height: 1.12 !important;
    }}

    .ps-kpi-value {{
        font-size: 22px !important;
        line-height: 1 !important;
        margin: 5px 0 !important;
        white-space: nowrap !important;
    }}

    .ps-kpi-sub {{
        font-size: 8.5px !important;
        line-height: 1.18 !important;
    }}

    .panel-title {{
        font-size: 14px !important;
        padding: 12px 13px !important;
        margin: 12px 0 8px 0 !important;
    }}

    .week-card-grid {{
        grid-template-columns: 1fr !important;
        gap: 10px !important;
        margin: 10px 0 16px 0 !important;
    }}

    .week-card-head {{
        font-size: 16px !important;
        padding: 10px 8px !important;
    }}

    .week-row {{
        padding: 9px 11px !important;
        gap: 8px !important;
    }}

    .week-row span {{
        font-size: 10px !important;
    }}

    .week-row b {{
        font-size: 16px !important;
    }}

    .week-row em {{
        font-size: 10px !important;
    }}

    /* Tablas: conservar estructura completa con desplazamiento horizontal */
    [data-testid="stDataFrame"],
    [data-testid="stDataEditor"],
    .ag-root-wrapper {{
        max-width: 100% !important;
        overflow-x: auto !important;
        -webkit-overflow-scrolling: touch !important;
    }}

    .ag-cell {{
        font-size: 10px !important;
    }}

    .ag-header-cell-text {{
        font-size: 10px !important;
    }}

    /* Gráficos adaptados al ancho del teléfono */
    [data-testid="stPlotlyChart"] {{
        width: 100% !important;
        overflow: hidden !important;
    }}

    [data-testid="stPlotlyChart"] > div {{
        width: 100% !important;
    }}

    .footer {{
        font-size: 10px !important;
        margin-top: 22px !important;
        padding-top: 12px !important;
    }}

    /* Evita que botones flotantes tapen contenido */
    .block-container {{
        padding-bottom: 5.5rem !important;
    }}
}}

/* Teléfonos muy angostos */
@media (max-width: 390px) {{
    .ps-kpi-grid {{
        gap: 7px !important;
    }}

    .ps-kpi-card {{
        padding: 9px 7px !important;
        gap: 6px !important;
    }}

    .ps-kpi-icon {{
        width: 42px !important;
        height: 42px !important;
        min-width: 42px !important;
        font-size: 20px !important;
    }}

    .ps-kpi-title {{
        font-size: 10px !important;
    }}

    .ps-kpi-value {{
        font-size: 19px !important;
    }}

    .ps-kpi-sub {{
        font-size: 7.7px !important;
    }}
}}


/* =========================================================
   V11 — DISEÑO RESPONSIVE EJECUTIVO
   ========================================================= */

:root {{
    --ps-gap: 14px;
    --ps-radius: 14px;
}}

.block-container {{
    max-width: 100% !important;
    width: 100% !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
}}

.ps-kpi-grid {{
    width: 100% !important;
    gap: var(--ps-gap) !important;
}}

.ps-kpi-card {{
    min-width: 0 !important;
    width: 100% !important;
    border-radius: var(--ps-radius) !important;
}}

.week-card-grid {{
    display: grid !important;
    grid-template-columns: repeat(4, minmax(0, 1fr)) !important;
    gap: 14px !important;
    width: 100% !important;
    margin: 12px 0 18px 0 !important;
}}

.week-card {{
    min-width: 0 !important;
    width: 100% !important;
    border-radius: 12px !important;
    overflow: hidden !important;
}}

.week-card-head {{
    padding: 11px 8px !important;
    font-size: 17px !important;
    line-height: 1.1 !important;
}}

.week-row {{
    display: grid !important;
    grid-template-columns: minmax(92px, 1fr) auto auto !important;
    align-items: center !important;
    gap: 8px !important;
    padding: 9px 12px !important;
    min-height: 43px !important;
}}

.week-row span {{
    font-size: 10px !important;
    white-space: nowrap !important;
}}

.week-row b {{
    font-size: 17px !important;
    white-space: nowrap !important;
}}

.week-row em {{
    min-width: 56px !important;
    text-align: right !important;
    font-size: 10px !important;
    white-space: nowrap !important;
}}

.panel-title {{
    margin-bottom: 6px !important;
}}

[data-testid="stDataFrame"],
[data-testid="stDataEditor"],
.ag-root-wrapper {{
    width: 100% !important;
    max-width: 100% !important;
    border-radius: 10px !important;
}}

[data-testid="stPlotlyChart"] {{
    width: 100% !important;
    max-width: 100% !important;
    margin-left: 0 !important;
    margin-right: 0 !important;
}}

[data-testid="stPlotlyChart"] > div,
.js-plotly-plot,
.plot-container,
.svg-container {{
    width: 100% !important;
    max-width: 100% !important;
}}

@media (min-width: 1440px) {{
    .block-container {{
        padding-left: 1.25rem !important;
        padding-right: 1.25rem !important;
    }}

    .ps-kpi-card {{
        min-height: 126px !important;
    }}

    .week-card-grid {{
        gap: 18px !important;
    }}

    .week-row {{
        min-height: 46px !important;
    }}
}}

@media (min-width: 769px) and (max-width: 1100px) {{
    .week-card-grid {{
        grid-template-columns: repeat(2, minmax(0, 1fr)) !important;
    }}

    .ps-kpi-grid {{
        grid-template-columns: repeat(3, minmax(0, 1fr)) !important;
    }}
}}

@media (max-width: 768px) {{
    .block-container {{
        padding-left: .5rem !important;
        padding-right: .5rem !important;
        padding-top: .3rem !important;
    }}

    .ps-kpi-grid {{
        grid-template-columns: repeat(2, minmax(0, 1fr)) !important;
        gap: 8px !important;
    }}

    .ps-kpi-card {{
        min-height: 104px !important;
        padding: 10px 8px !important;
        gap: 7px !important;
        border-radius: 12px !important;
    }}

    .ps-kpi-card:nth-child(5) {{
        grid-column: 1 / -1 !important;
        min-height: 90px !important;
    }}

    .ps-kpi-icon {{
        width: 44px !important;
        height: 44px !important;
        min-width: 44px !important;
        font-size: 21px !important;
    }}

    .ps-kpi-title {{
        font-size: 10px !important;
        line-height: 1.1 !important;
    }}

    .ps-kpi-value {{
        font-size: 20px !important;
        line-height: 1 !important;
        margin: 4px 0 !important;
    }}

    .ps-kpi-sub {{
        font-size: 7.8px !important;
        line-height: 1.15 !important;
    }}

    .week-card-grid {{
        grid-template-columns: repeat(2, minmax(0, 1fr)) !important;
        gap: 8px !important;
        margin: 10px 0 14px 0 !important;
    }}

    .week-card-head {{
        padding: 9px 6px !important;
        font-size: 14px !important;
    }}

    .week-row {{
        grid-template-columns: 1fr auto !important;
        gap: 5px !important;
        padding: 7px 8px !important;
        min-height: 38px !important;
    }}

    .week-row span {{
        font-size: 8px !important;
    }}

    .week-row b {{
        font-size: 13px !important;
    }}

    .week-row em {{
        grid-column: 1 / -1 !important;
        min-width: 0 !important;
        text-align: right !important;
        font-size: 8px !important;
        margin-top: -3px !important;
    }}

    [data-testid="stDataFrame"],
    [data-testid="stDataEditor"],
    .ag-root-wrapper {{
        overflow-x: auto !important;
        -webkit-overflow-scrolling: touch !important;
    }}

    .ag-header-cell-text,
    .ag-cell {{
        font-size: 9px !important;
    }}

    [data-testid="stPlotlyChart"] {{
        width: calc(100vw - 1rem) !important;
        margin-left: 0 !important;
        margin-right: 0 !important;
    }}

    [data-testid="stPlotlyChart"] .svg-container {{
        min-height: 430px !important;
    }}
}}

@media (max-width: 390px) {{
    .ps-kpi-grid {{
        gap: 6px !important;
    }}

    .ps-kpi-card {{
        padding: 8px 6px !important;
    }}

    .ps-kpi-icon {{
        width: 40px !important;
        height: 40px !important;
        min-width: 40px !important;
    }}

    .ps-kpi-value {{
        font-size: 18px !important;
    }}

    .week-card-grid {{
        gap: 6px !important;
    }}

    .week-row {{
        padding: 6px 6px !important;
    }}
}}


/* =========================================================
   V11.3 — MENÚ TIPO CARRUSEL
   Inspirado en navegación móvil por tarjetas deslizables
   ========================================================= */

.ps-mobile-nav-title {{
    display:none;
}}

/* Escritorio: navegación corporativa horizontal */
@media (min-width: 769px) {{
    .ps-tabbar {{
        background:var(--azul)!important;
        border-top:5px solid var(--rosa)!important;
        overflow-x:auto!important;
        scroll-behavior:smooth!important;
    }}

    .ps-tabbar [role="radiogroup"] {{
        display:flex!important;
        flex-wrap:nowrap!important;
        min-width:max-content!important;
    }}

    .ps-tabbar label {{
        min-height:58px!important;
        padding:0 20px!important;
        background:var(--azul)!important;
        color:rgba(255,255,255,.76)!important;
        transition:background .18s ease, color .18s ease!important;
    }}

    .ps-tabbar label:has(input:checked) {{
        background:#142E73!important;
        box-shadow:inset 0 -5px 0 var(--rosa)!important;
        color:#FFF!important;
    }}
}}

/* Móvil: carrusel de reportes */
@media (max-width: 768px) {{
    .ps-mobile-nav-title {{
        display:flex!important;
        align-items:center!important;
        justify-content:space-between!important;
        margin:8px 2px 7px 2px!important;
        color:#5B6476!important;
        font-size:11px!important;
        font-weight:800!important;
    }}

    .ps-mobile-nav-arrow {{
        color:var(--rosa)!important;
        font-size:17px!important;
        font-weight:900!important;
    }}

    .ps-tabbar {{
        position:relative!important;
        left:auto!important;
        right:auto!important;
        width:calc(100% + 1rem)!important;
        max-width:none!important;
        margin:0 -.5rem 16px -.5rem!important;
        padding:8px 0 12px 0!important;
        background:
            linear-gradient(135deg,#111A55 0%,#24126E 55%,#3B146E 100%)!important;
        border-top:4px solid var(--rosa)!important;
        border-bottom:1px solid rgba(255,255,255,.12)!important;
        overflow-x:auto!important;
        overflow-y:hidden!important;
        scroll-snap-type:x mandatory!important;
        scroll-padding-inline:calc(50vw - 92px)!important;
        -webkit-overflow-scrolling:touch!important;
        scrollbar-width:none!important;
    }}

    .ps-tabbar::-webkit-scrollbar {{
        display:none!important;
    }}

    .ps-tabbar [role="radiogroup"] {{
        display:flex!important;
        flex-wrap:nowrap!important;
        align-items:center!important;
        gap:10px!important;
        width:max-content!important;
        min-width:max-content!important;
        padding:0 calc(50vw - 92px)!important;
        min-height:106px!important;
    }}

    .ps-tabbar label {{
        position:relative!important;
        flex:0 0 154px!important;
        width:154px!important;
        min-width:154px!important;
        height:82px!important;
        min-height:82px!important;
        padding:39px 9px 8px 9px!important;
        border-radius:15px!important;
        border:1px solid rgba(255,255,255,.22)!important;
        background:rgba(255,255,255,.10)!important;
        color:rgba(255,255,255,.82)!important;
        display:flex!important;
        align-items:center!important;
        justify-content:center!important;
        text-align:center!important;
        white-space:normal!important;
        scroll-snap-align:center!important;
        box-shadow:0 8px 18px rgba(0,0,0,.16)!important;
        transform:scale(.91)!important;
        opacity:.76!important;
        transition:
            transform .22s ease,
            opacity .22s ease,
            background .22s ease,
            border-color .22s ease!important;
    }}

    .ps-tabbar label p,
    .ps-tabbar label span {{
        color:inherit!important;
        font-size:11px!important;
        line-height:1.08!important;
        font-weight:900!important;
        text-align:center!important;
    }}

    .ps-tabbar label::before {{
        content:"▦";
        position:absolute!important;
        top:9px!important;
        left:50%!important;
        transform:translateX(-50%)!important;
        width:27px!important;
        height:27px!important;
        border-radius:50%!important;
        display:flex!important;
        align-items:center!important;
        justify-content:center!important;
        background:rgba(255,255,255,.16)!important;
        color:#FFF!important;
        font-size:15px!important;
        font-weight:900!important;
    }}

    /* Iconos por reporte */
    .ps-tabbar label:nth-child(1)::before {{content:"▦";}}
    .ps-tabbar label:nth-child(2)::before {{content:"◷";}}
    .ps-tabbar label:nth-child(3)::before {{content:"W";}}
    .ps-tabbar label:nth-child(4)::before {{content:"M";}}
    .ps-tabbar label:nth-child(5)::before {{content:"↗";}}
    .ps-tabbar label:nth-child(6)::before {{content:"$";}}
    .ps-tabbar label:nth-child(7)::before {{content:"✓";}}
    .ps-tabbar label:nth-child(8)::before {{content:"↻";}}
    .ps-tabbar label:nth-child(9)::before {{content:"#";}}
    .ps-tabbar label:nth-child(10)::before {{content:"Σ";}}
    .ps-tabbar label:nth-child(11)::before {{content:"!";}}
    .ps-tabbar label:nth-child(12)::before {{content:"⚙";}}
    .ps-tabbar label:nth-child(13)::before {{content:"♙";}}

    .ps-tabbar label:hover {{
        background:rgba(255,255,255,.15)!important;
        color:#FFF!important;
    }}

    .ps-tabbar label:has(input:checked) {{
        transform:scale(1.04)!important;
        opacity:1!important;
        z-index:3!important;
        background:
            linear-gradient(145deg,rgba(255,255,255,.24),rgba(255,255,255,.13))!important;
        border:2px solid #FFF!important;
        box-shadow:
            0 12px 25px rgba(0,0,0,.25),
            0 0 0 3px rgba(255,0,128,.30)!important;
        color:#FFF!important;
    }}

    .ps-tabbar label:has(input:checked)::before {{
        background:var(--rosa)!important;
        box-shadow:0 4px 10px rgba(255,0,128,.35)!important;
    }}

    .ps-tabbar label:has(input:checked)::after {{
        content:"";
        position:absolute!important;
        bottom:-9px!important;
        left:50%!important;
        transform:translateX(-50%)!important;
        width:30px!important;
        height:4px!important;
        border-radius:4px!important;
        background:var(--rosa)!important;
    }}

    .ps-tabbar input[type="radio"] {{
        position:absolute!important;
        opacity:0!important;
        pointer-events:none!important;
    }}
}}

/* Teléfono angosto */
@media (max-width: 390px) {{
    .ps-tabbar [role="radiogroup"] {{
        padding-left:calc(50vw - 82px)!important;
        padding-right:calc(50vw - 82px)!important;
        gap:8px!important;
    }}

    .ps-tabbar label {{
        flex-basis:140px!important;
        width:140px!important;
        min-width:140px!important;
        height:78px!important;
        min-height:78px!important;
    }}
}}


/* V11.4 — CARRUSEL HORIZONTAL REAL */
.st-key-nav_v114_carousel,
.st-key-nav_v113_carousel {{
    width: calc(100% + 3.2rem) !important;
    margin-left: -1.6rem !important;
    margin-right: -1.6rem !important;
    margin-bottom: 22px !important;
    padding: 0 1.6rem !important;
    box-sizing: border-box !important;
    background: var(--azul) !important;
    border-top: 4px solid var(--rosa) !important;
    overflow: hidden !important;
}}
.st-key-nav_v114_carousel [data-testid="stRadio"],
.st-key-nav_v113_carousel [data-testid="stRadio"] {{
    width: 100% !important;
}}
.st-key-nav_v114_carousel [role="radiogroup"],
.st-key-nav_v113_carousel [role="radiogroup"] {{
    display: flex !important;
    flex-flow: row nowrap !important;
    align-items: stretch !important;
    gap: 0 !important;
    width: max-content !important;
    min-width: 100% !important;
    min-height: 58px !important;
    overflow: visible !important;
}}
.st-key-nav_v114_carousel label,
.st-key-nav_v113_carousel label {{
    flex: 0 0 auto !important;
    min-width: max-content !important;
    min-height: 58px !important;
    padding: 0 18px !important;
    margin: 0 !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    border: 0 !important;
    border-radius: 0 !important;
    background: transparent !important;
    color: rgba(255,255,255,.76) !important;
    white-space: nowrap !important;
    font-weight: 850 !important;
    box-shadow: none !important;
}}
.st-key-nav_v114_carousel label:hover,
.st-key-nav_v113_carousel label:hover {{
    color: #FFFFFF !important;
    background: rgba(255,255,255,.06) !important;
}}
.st-key-nav_v114_carousel label:has(input:checked),
.st-key-nav_v113_carousel label:has(input:checked) {{
    color: #FFFFFF !important;
    background: #142E73 !important;
    box-shadow: inset 0 -5px 0 var(--rosa) !important;
}}
.st-key-nav_v114_carousel label *,
.st-key-nav_v113_carousel label * {{
    color: inherit !important;
    font-weight: inherit !important;
}}
.st-key-nav_v114_carousel [data-testid="stRadio"] input,
.st-key-nav_v113_carousel [data-testid="stRadio"] input,
.st-key-nav_v114_carousel [data-testid="stRadio"] [data-baseweb="radio"] > div:first-child,
.st-key-nav_v113_carousel [data-testid="stRadio"] [data-baseweb="radio"] > div:first-child {{
    position: absolute !important;
    opacity: 0 !important;
    width: 1px !important;
    height: 1px !important;
    pointer-events: none !important;
}}
@media (max-width: 768px) {{
    .ps-mobile-nav-title {{
        display: flex !important;
        align-items: center !important;
        justify-content: space-between !important;
        margin: 8px 2px 7px 2px !important;
        color: #5B6476 !important;
        font-size: 12px !important;
        font-weight: 850 !important;
    }}
    .ps-mobile-nav-arrow {{
        color: var(--rosa) !important;
        font-size: 18px !important;
        font-weight: 900 !important;
    }}
    .st-key-nav_v114_carousel,
    .st-key-nav_v113_carousel {{
        width: calc(100% + 1rem) !important;
        margin-left: -.5rem !important;
        margin-right: -.5rem !important;
        margin-bottom: 18px !important;
        padding: 10px 0 12px 0 !important;
        background: linear-gradient(135deg,#111A55 0%,#24126E 55%,#3B146E 100%) !important;
        border-top: 4px solid var(--rosa) !important;
        border-bottom: 1px solid rgba(255,255,255,.14) !important;
        overflow-x: auto !important;
        overflow-y: hidden !important;
        scroll-snap-type: x proximity !important;
        scroll-padding-inline: 18px !important;
        overscroll-behavior-x: contain !important;
        -webkit-overflow-scrolling: touch !important;
        scrollbar-width: none !important;
        touch-action: pan-x !important;
    }}
    .st-key-nav_v114_carousel::-webkit-scrollbar,
    .st-key-nav_v113_carousel::-webkit-scrollbar {{
        display: none !important;
    }}
    .st-key-nav_v114_carousel [data-testid="stRadio"],
    .st-key-nav_v113_carousel [data-testid="stRadio"] {{
        width: max-content !important;
        min-width: max-content !important;
        overflow: visible !important;
    }}
    .st-key-nav_v114_carousel [role="radiogroup"],
    .st-key-nav_v113_carousel [role="radiogroup"] {{
        display: flex !important;
        flex-flow: row nowrap !important;
        width: max-content !important;
        min-width: max-content !important;
        gap: 10px !important;
        padding: 0 14px !important;
        min-height: 58px !important;
    }}
    .st-key-nav_v114_carousel label,
    .st-key-nav_v113_carousel label {{
        flex: 0 0 auto !important;
        width: auto !important;
        min-width: 126px !important;
        max-width: 190px !important;
        height: 54px !important;
        min-height: 54px !important;
        padding: 0 15px !important;
        border-radius: 12px !important;
        border: 1px solid rgba(255,255,255,.22) !important;
        background: rgba(255,255,255,.09) !important;
        color: rgba(255,255,255,.80) !important;
        white-space: nowrap !important;
        text-align: center !important;
        scroll-snap-align: center !important;
        box-shadow: 0 5px 13px rgba(0,0,0,.14) !important;
        transform: none !important;
        opacity: 1 !important;
        font-size: 13px !important;
        line-height: 1.1 !important;
    }}
    .st-key-nav_v114_carousel label:has(input:checked),
    .st-key-nav_v113_carousel label:has(input:checked) {{
        background: #FFFFFF !important;
        color: var(--azul) !important;
        border-color: #FFFFFF !important;
        box-shadow: 0 7px 18px rgba(0,0,0,.22), inset 0 -5px 0 var(--rosa) !important;
    }}
    .st-key-nav_v114_carousel label:has(input:checked) *,
    .st-key-nav_v113_carousel label:has(input:checked) * {{
        color: var(--azul) !important;
    }}
}}
@media (min-width: 769px) {{
    .ps-mobile-nav-title {{
        display: none !important;
    }}
}}


/* V11.5: el carrusel es un componente independiente; ocultar restos del menú radio anterior. */
.ps-mobile-nav-title,
.st-key-nav_v114_carousel,
.st-key-nav_v113_carousel {{
    display: none !important;
}}


/* V11.6 — CARRUSEL CLICABLE SIN IFRAME */
.ps-carousel-title {{ display:none; }}
.ps-carousel-shell {{
  width:calc(100% + 3.2rem); margin-left:-1.6rem; margin-right:-1.6rem; margin-bottom:22px;
  background:linear-gradient(135deg,#111A55 0%,#24126E 55%,#3B146E 100%);
  border-top:4px solid var(--rosa); border-bottom:1px solid rgba(255,255,255,.14);
  overflow-x:auto; overflow-y:hidden; -webkit-overflow-scrolling:touch; scrollbar-width:none;
  overscroll-behavior-x:contain; touch-action:pan-x;
}}
.ps-carousel-shell::-webkit-scrollbar {{ display:none; }}
.ps-carousel-track {{ display:flex; flex-flow:row nowrap; align-items:stretch; gap:0; width:max-content; min-width:100%; min-height:58px; padding:0 1.6rem; }}
.ps-carousel-card {{
  flex:0 0 auto; min-width:max-content; min-height:58px; padding:0 18px; display:flex; align-items:center; justify-content:center;
  color:rgba(255,255,255,.76)!important; text-decoration:none!important; font-weight:850; white-space:nowrap;
  background:transparent; border-radius:0; box-shadow:none; cursor:pointer; -webkit-tap-highlight-color:transparent; user-select:none;
}}
.ps-carousel-card:hover {{ color:#fff!important; background:rgba(255,255,255,.06); }}
.ps-carousel-card.active {{ color:#fff!important; background:#142E73; box-shadow:inset 0 -5px 0 var(--rosa); }}
@media (max-width:768px) {{
  .ps-carousel-title {{ display:flex; align-items:center; justify-content:space-between; margin:8px 2px 7px; color:#5B6476; font-size:12px; font-weight:850; }}
  .ps-carousel-arrow {{ color:var(--rosa); font-size:18px; font-weight:900; }}
  .ps-carousel-shell {{ width:calc(100% + 1rem); margin-left:-.5rem; margin-right:-.5rem; margin-bottom:18px; padding:10px 0 12px; scroll-snap-type:x proximity; scroll-padding-inline:18px; }}
  .ps-carousel-track {{ gap:10px; min-width:max-content; padding:0 14px; min-height:58px; }}
  .ps-carousel-card {{
    flex:0 0 154px; width:154px; min-width:154px; max-width:190px; height:58px; min-height:58px; padding:0 15px;
    border-radius:14px; border:1px solid rgba(255,255,255,.22); background:rgba(255,255,255,.09);
    color:rgba(255,255,255,.80)!important; text-align:center; white-space:normal; line-height:1.08; font-size:13px;
    scroll-snap-align:center; box-shadow:0 5px 13px rgba(0,0,0,.14); transform:scale(.92); opacity:.78;
  }}
  .ps-carousel-card.active {{
    background:#fff; color:var(--azul)!important; border-color:#fff;
    box-shadow:0 7px 18px rgba(0,0,0,.22), inset 0 -5px 0 var(--rosa); transform:scale(1); opacity:1;
  }}
}}


/* V11.7 — MENÚ NATIVO SIN PÉRDIDA DE SESIÓN */
.ps-carousel-title {{
    display:none;
}}
.st-key-nav_session_safe {{
    width: calc(100% + 3.2rem) !important;
    margin-left: -1.6rem !important;
    margin-right: -1.6rem !important;
    margin-bottom: 22px !important;
    padding: 0 1.6rem !important;
    background: var(--azul) !important;
    border-top: 4px solid var(--rosa) !important;
    overflow-x: auto !important;
    overflow-y: hidden !important;
    -webkit-overflow-scrolling: touch !important;
    scrollbar-width: none !important;
}}
.st-key-nav_session_safe::-webkit-scrollbar {{ display:none !important; }}
.st-key-nav_session_safe [role="radiogroup"] {{
    display:flex !important;
    flex-flow:row nowrap !important;
    width:max-content !important;
    min-width:100% !important;
    gap:0 !important;
}}
.st-key-nav_session_safe label {{
    flex:0 0 auto !important;
    min-width:max-content !important;
    min-height:58px !important;
    padding:0 18px !important;
    display:flex !important;
    align-items:center !important;
    justify-content:center !important;
    white-space:nowrap !important;
    color:rgba(255,255,255,.78) !important;
    background:transparent !important;
    border-radius:0 !important;
    font-weight:850 !important;
}}
.st-key-nav_session_safe label:has(input:checked) {{
    color:#fff !important;
    background:#142E73 !important;
    box-shadow:inset 0 -5px 0 var(--rosa) !important;
}}
.st-key-nav_session_safe label * {{ color:inherit !important; font-weight:inherit !important; }}
.st-key-nav_session_safe input,
.st-key-nav_session_safe [data-baseweb="radio"] > div:first-child {{
    position:absolute !important;
    opacity:0 !important;
    width:1px !important;
    height:1px !important;
}}

/* PORTAL DE ACCESO ESTILO PRICE SHOES */
[data-testid="stSidebar"]:has(.login-portal-shell) {{ display:none !important; }}
.login-portal-shell {{
    position:relative;
    max-width:760px;
    min-height:290px;
    margin:50px auto 0;
    border-radius:24px 24px 0 0;
    overflow:hidden;
    background:
      linear-gradient(rgba(3,25,20,.78),rgba(3,25,20,.86)),
      radial-gradient(circle at 50% 15%,rgba(236,0,126,.20),transparent 38%),
      linear-gradient(135deg,#10245F,#063A2D);
    box-shadow:0 20px 55px rgba(16,36,95,.20);
}}
.login-portal-brand {{
    position:relative;
    z-index:2;
    display:flex;
    flex-direction:column;
    align-items:center;
    padding:44px 20px 28px;
    color:#fff;
}}
.login-portal-logo {{
    width:150px;
    height:92px;
    border:3px solid #fff;
    border-radius:50%;
    display:flex;
    align-items:center;
    justify-content:center;
    text-align:center;
    font-family:Georgia,serif;
    font-size:34px;
    line-height:.72;
    font-weight:900;
    text-shadow:0 2px 5px rgba(0,0,0,.45);
}}
.login-portal-title {{ margin-top:18px; font-size:30px; font-weight:900; }}
.login-portal-subtitle {{ margin-top:5px; font-size:13px; opacity:.8; }}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) {{
    max-width:760px;
    margin:0 auto 40px;
    padding:0 150px 36px;
    background:#08251F;
    border-radius:0 0 24px 24px;
    border:0 !important;
    box-shadow:0 20px 55px rgba(16,36,95,.20);
}}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) label {{ color:#fff !important; }}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) input {{
    color:#fff !important;
    background:transparent !important;
    border:0 !important;
    border-bottom:1px solid rgba(255,255,255,.7) !important;
    border-radius:0 !important;
}}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) button {{
    margin-top:12px !important;
    background:var(--rosa) !important;
    color:#fff !important;
    border-radius:0 !important;
    min-height:52px !important;
    font-weight:900 !important;
}}
@media(max-width:768px) {{
    .ps-carousel-title {{
        display:flex;
        align-items:center;
        justify-content:space-between;
        margin:8px 2px 7px;
        color:#5B6476;
        font-size:12px;
        font-weight:850;
    }}
    .ps-carousel-arrow {{ color:var(--rosa); font-size:18px; }}
    .st-key-nav_session_safe {{
        width:calc(100% + 1rem) !important;
        margin-left:-.5rem !important;
        margin-right:-.5rem !important;
        padding:10px 0 12px !important;
        background:linear-gradient(135deg,#111A55,#24126E 55%,#3B146E) !important;
        scroll-snap-type:x proximity !important;
    }}
    .st-key-nav_session_safe [role="radiogroup"] {{ gap:10px !important; padding:0 14px !important; min-width:max-content !important; }}
    .st-key-nav_session_safe label {{
        min-width:154px !important;
        width:154px !important;
        min-height:58px !important;
        border:1px solid rgba(255,255,255,.22) !important;
        border-radius:14px !important;
        background:rgba(255,255,255,.09) !important;
        color:rgba(255,255,255,.8) !important;
        scroll-snap-align:center !important;
    }}
    .st-key-nav_session_safe label:has(input:checked) {{
        background:#fff !important;
        color:var(--azul) !important;
        border-color:#fff !important;
        box-shadow:0 7px 18px rgba(0,0,0,.22), inset 0 -5px 0 var(--rosa) !important;
    }}
    .login-portal-shell {{ margin:20px auto 0; min-height:230px; border-radius:18px 18px 0 0; }}
    .login-portal-brand {{ padding:28px 14px 22px; }}
    .login-portal-logo {{ width:118px; height:72px; font-size:27px; }}
    .login-portal-title {{ font-size:25px; }}
    [data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) {{
        padding:0 24px 28px;
        border-radius:0 0 18px 18px;
    }}
}}


/* V11.8 — Portal corporativo, login completo y menú superior */
.login-fullscreen-bg {{
    position: fixed;
    inset: 0;
    z-index: -1;
    background:
        linear-gradient(rgba(0,25,28,.82), rgba(0,36,31,.92)),
        radial-gradient(circle at 50% 20%, rgba(255,255,255,.08), transparent 38%),
        linear-gradient(145deg,#071F2A,#003B31);
}}
.login-brand-zone {{
    max-width: 650px;
    margin: 5vh auto 0;
    text-align: center;
    color: #fff;
}}
.login-portal-title {{
    font-size: 40px !important;
    margin-top: 20px;
}}
.login-portal-subtitle {{
    font-size: 25px !important;
    font-weight: 800;
    margin-top: 4px;
}}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) {{
    max-width: 650px !important;
    margin: 25px auto 0 !important;
    padding: 28px 38px 34px !important;
    border-radius: 18px !important;
    background: rgba(0,39,34,.90) !important;
    border: 1px solid rgba(255,255,255,.18) !important;
    box-shadow: 0 18px 50px rgba(0,0,0,.30) !important;
}}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) label,
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) p {{
    color: #fff !important;
    font-weight: 800 !important;
}}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) input {{
    color: #111827 !important;
    background: #fff !important;
    -webkit-text-fill-color: #111827 !important;
}}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) input::placeholder {{
    color: #667085 !important;
    opacity: 1 !important;
}}
.portal-main-brand {{
    display: flex;
    align-items: center;
    gap: 18px;
    min-height: 88px;
}}
.portal-main-logo {{
    width: 110px;
    min-width: 110px;
}}
.portal-main-title {{
    color: var(--azul);
    font-weight: 900;
    font-size: 31px;
    line-height: 1.05;
}}
.portal-main-subtitle {{
    color: #596174;
    font-weight: 700;
    font-size: 14px;
    margin-top: 8px;
}}
.portal-user-date {{
    text-align: right;
    color: #71798A;
    font-size: 11px;
    margin-top: 4px;
}}
.portal-pink-line {{
    height: 5px;
    background: var(--rosa);
    margin: 2px -1rem 14px;
}}
.portal-readonly-badge {{
    border: 1px solid #D5DCEA;
    background: #fff;
    border-radius: 8px;
    padding: 10px 14px;
    color: #596174;
    text-align: center;
    font-weight: 750;
}}
@media (max-width: 768px) {{
    .login-brand-zone {{
        margin-top: 2vh;
        padding: 0 14px;
    }}
    .login-portal-title {{
        font-size: 28px !important;
    }}
    .login-portal-subtitle {{
        font-size: 19px !important;
    }}
    [data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) {{
        margin: 18px 10px 0 !important;
        padding: 22px 18px 26px !important;
    }}
    .portal-main-brand {{
        gap: 10px;
        min-height: 66px;
    }}
    .portal-main-logo {{
        width: 72px;
        min-width: 72px;
    }}
    .portal-main-title {{
        font-size: 21px;
    }}
    .portal-main-subtitle {{
        font-size: 10px;
        margin-top: 4px;
    }}
}}


/* V11.9 — corrección de acceso, cabecera y administración */
.login-brand-card {{
    text-align: center;
    color: #fff;
    margin: 0 auto 18px;
}}
.login-real-logo {{
    width: 180px;
    margin: 0 auto 8px;
}}
.login-real-logo img {{
    width: 100% !important;
    max-height: 120px !important;
    object-fit: contain !important;
    filter: brightness(0) invert(1) !important;
}}
.login-portal-title {{
    color: #fff !important;
    font-size: 40px !important;
    font-weight: 900 !important;
    line-height: 1.05 !important;
    margin-top: 8px !important;
}}
.login-portal-subtitle {{
    color: #fff !important;
    font-size: 24px !important;
    font-weight: 800 !important;
    margin-top: 4px !important;
}}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) {{
    width: 100% !important;
    max-width: 760px !important;
    margin: 0 auto !important;
    padding: 28px 38px 34px !important;
    border-radius: 18px !important;
    background: rgba(0,55,47,.94) !important;
    border: 1px solid rgba(255,255,255,.20) !important;
    box-shadow: 0 20px 55px rgba(0,0,0,.35) !important;
}}
[data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) input {{
    color: #111827 !important;
    -webkit-text-fill-color: #111827 !important;
    background: #fff !important;
}}
.portal-header-spacer {{
    height: 28px;
}}
.portal-main-brand {{
    padding-top: 12px !important;
    padding-bottom: 12px !important;
    overflow: visible !important;
}}
.portal-main-logo {{
    background: transparent !important;
    border: 0 !important;
    overflow: visible !important;
}}
.portal-main-logo img {{
    background: transparent !important;
    mix-blend-mode: multiply !important;
    object-fit: contain !important;
}}
.portal-main-title,
.portal-main-subtitle {{
    overflow: visible !important;
}}
[data-testid="stPopover"] button {{
    min-height: 46px !important;
}}
@media (max-width:768px) {{
    .login-real-logo {{
        width: 135px;
    }}
    .login-portal-title {{
        font-size: 30px !important;
    }}
    .login-portal-subtitle {{
        font-size: 20px !important;
    }}
    [data-testid="stForm"]:has(input[aria-label="Usuario o correo"]) {{
        padding: 22px 18px 26px !important;
    }}
    .portal-header-spacer {{
        height: 16px;
    }}
}}


/* V12 — Portal de aplicaciones y módulo Cambios y Muertos */
.portal-home-brand {{
    display:flex;
    align-items:center;
    gap:18px;
    min-height:92px;
    padding:12px 0;
}}
.portal-home-logo {{
    width:105px;
    min-width:105px;
}}
.portal-home-title {{
    color:var(--azul);
    font-size:32px;
    font-weight:900;
    line-height:1.05;
}}
.portal-home-subtitle {{
    color:#667085;
    font-size:14px;
    font-weight:700;
    margin-top:7px;
}}
.portal-section-title {{
    margin-top:22px;
    color:var(--azul);
    font-size:28px;
    font-weight:900;
}}
.portal-section-subtitle {{
    color:#6B7280;
    margin:5px 0 18px;
}}
.app-tile {{
    display:flex;
    align-items:center;
    gap:20px;
    min-height:150px;
    padding:24px 28px;
    border:1px solid #D9E1EF;
    border-radius:16px;
    background:#FFFFFF;
    box-shadow:0 12px 30px rgba(31,42,68,.08);
}}
.app-tile-icon {{
    display:flex;
    align-items:center;
    justify-content:center;
    width:72px;
    height:72px;
    min-width:72px;
    border-radius:18px;
    background:linear-gradient(135deg,var(--azul),#402080);
    color:#FFFFFF;
    font-size:36px;
    font-weight:900;
}}
.app-tile-title {{
    color:var(--azul);
    font-size:25px;
    font-weight:900;
}}
.app-tile-subtitle {{
    color:#667085;
    margin-top:6px;
    line-height:1.35;
}}

/* Pestañas internas de Cambios y Muertos */
.st-key-nav_v120_tabs {{
    width:calc(100% + 2rem) !important;
    margin-left:-1rem !important;
    margin-right:-1rem !important;
    margin-bottom:22px !important;
    padding:0 1rem !important;
    background:var(--azul) !important;
    border-top:4px solid var(--rosa) !important;
    overflow-x:auto !important;
    overflow-y:hidden !important;
    scrollbar-width:none !important;
    -webkit-overflow-scrolling:touch !important;
}}
.st-key-nav_v120_tabs::-webkit-scrollbar {{
    display:none !important;
}}
.st-key-nav_v120_tabs [role="radiogroup"] {{
    display:flex !important;
    flex-flow:row nowrap !important;
    width:max-content !important;
    min-width:100% !important;
    gap:0 !important;
}}
.st-key-nav_v120_tabs label {{
    flex:0 0 auto !important;
    min-width:max-content !important;
    min-height:58px !important;
    padding:0 18px !important;
    margin:0 !important;
    display:flex !important;
    align-items:center !important;
    justify-content:center !important;
    color:rgba(255,255,255,.76) !important;
    white-space:nowrap !important;
    font-weight:850 !important;
    background:transparent !important;
}}
.st-key-nav_v120_tabs label:has(input:checked) {{
    color:#FFFFFF !important;
    background:#142E73 !important;
    box-shadow:inset 0 -5px 0 var(--rosa) !important;
}}
.st-key-nav_v120_tabs label *,
.st-key-nav_v120_tabs label:has(input:checked) * {{
    color:inherit !important;
    font-weight:inherit !important;
}}
.st-key-nav_v120_tabs input,
.st-key-nav_v120_tabs [data-baseweb="radio"] > div:first-child {{
    position:absolute !important;
    opacity:0 !important;
    width:1px !important;
    height:1px !important;
}}

@media(max-width:768px) {{
    .portal-home-brand {{
        gap:10px;
        min-height:70px;
    }}
    .portal-home-logo {{
        width:72px;
        min-width:72px;
    }}
    .portal-home-title {{
        font-size:23px;
    }}
    .portal-home-subtitle {{
        font-size:10px;
    }}
    .app-tile {{
        padding:18px;
        min-height:125px;
    }}
    .app-tile-icon {{
        width:58px;
        height:58px;
        min-width:58px;
        font-size:28px;
    }}
    .app-tile-title {{
        font-size:20px;
    }}
    .st-key-nav_v120_tabs {{
        width:calc(100% + 1rem) !important;
        margin-left:-.5rem !important;
        margin-right:-.5rem !important;
        padding:10px 0 12px !important;
        background:linear-gradient(135deg,#111A55,#24126E 55%,#3B146E) !important;
        scroll-snap-type:x proximity !important;
    }}
    .st-key-nav_v120_tabs [role="radiogroup"] {{
        gap:10px !important;
        padding:0 14px !important;
        min-width:max-content !important;
    }}
    .st-key-nav_v120_tabs label {{
        min-width:145px !important;
        height:56px !important;
        min-height:56px !important;
        border-radius:13px !important;
        border:1px solid rgba(255,255,255,.22) !important;
        background:rgba(255,255,255,.09) !important;
        scroll-snap-align:center !important;
    }}
    .st-key-nav_v120_tabs label:has(input:checked) {{
        background:#FFFFFF !important;
        color:var(--azul) !important;
        border-color:#FFFFFF !important;
        box-shadow:0 7px 18px rgba(0,0,0,.22), inset 0 -5px 0 var(--rosa) !important;
    }}
}}

</style>
""",
        unsafe_allow_html=True,
    )


def render_portal_header():
    now = datetime.now(MX_TZ)
    user = st.session_state.get("user", {})
    user_name = user.get("nombre", "Consulta")
    permiso = user.get("permiso", "Consulta")
    nomina = user.get("nomina", "")

    st.markdown('<div class="portal-header-spacer"></div>', unsafe_allow_html=True)
    c_logo, c_user = st.columns([7.5, 2.5], vertical_alignment="center")

    with c_logo:
        st.markdown(
            f"""
            <div class="portal-home-brand">
                <div class="portal-home-logo">{logo_html()}</div>
                <div>
                    <div class="portal-home-title">Operaciones Ropa</div>
                    <div class="portal-home-subtitle">Portal de aplicaciones e indicadores</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with c_user:
        with st.popover(f"👤 {user_name}", use_container_width=True):
            st.markdown(f"**Usuario:** {user_name}")
            if nomina:
                st.caption(f"Nómina: {nomina}")
            st.caption(f"Permiso: {permiso}")
            st.caption(f"Fecha: {now.strftime('%d/%m/%Y')}")
            if st.button("Cerrar sesión", key="logout_portal", use_container_width=True):
                st.session_state.pop("user", None)
                st.session_state.pop("active_app", None)
                st.session_state.pop("nav_page", None)
                st.rerun()

        st.markdown(
            f'<div class="portal-user-date">{now.strftime("%d/%m/%Y")} · {permiso}</div>',
            unsafe_allow_html=True,
        )

    st.markdown('<div class="portal-pink-line"></div>', unsafe_allow_html=True)

def render_header():
    now = datetime.now(MX_TZ)
    user = st.session_state.get("user", {})
    user_name = user.get("nombre", "Consulta")
    permiso = user.get("permiso", "Consulta")

    st.markdown('<div class="portal-header-spacer"></div>', unsafe_allow_html=True)
    c_back, c_brand, c_user = st.columns([0.7, 6.8, 2.5], vertical_alignment="center")

    with c_back:
        if st.button("←", key="back_to_apps", help="Volver al menú principal", use_container_width=True):
            st.session_state["active_app"] = None
            st.session_state["nav_page"] = "Resumen"
            st.rerun()

    with c_brand:
        st.markdown(
            f"""
            <div class="portal-main-brand">
                <div class="portal-main-logo">{logo_html()}</div>
                <div class="portal-main-copy">
                    <div class="portal-main-title">Indicadores Cambios y Muertos</div>
                    <div class="portal-main-subtitle">Recuperación · Productividad · Conversión</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with c_user:
        with st.popover(f"👤 {user_name}", use_container_width=True):
            st.caption(f"Permiso: {permiso}")
            st.caption(f"Fecha: {now.strftime('%d/%m/%Y')}")
            if st.button("Cerrar sesión", key="logout_top", use_container_width=True):
                st.session_state.pop("user", None)
                st.session_state.pop("active_app", None)
                st.session_state.pop("nav_page", None)
                st.rerun()
        st.markdown(
            f'<div class="portal-user-date">{now.strftime("%d/%m/%Y")} · {permiso}</div>',
            unsafe_allow_html=True,
        )

    st.markdown('<div class="portal-pink-line"></div>', unsafe_allow_html=True)


def save_uploaded_file(uploaded):
    ACTIVE_FILE.write_bytes(uploaded.getbuffer())
    META_FILE.write_text(
        json.dumps(
            {
                "nombre_original": uploaded.name,
                "fecha_carga": datetime.now(MX_TZ).strftime("%Y-%m-%d %H:%M:%S"),
                "mtime": ACTIVE_FILE.stat().st_mtime,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    clear_cache_files()


def clear_cache_files():
    for p in CACHE_DIR.glob("*"):
        try:
            p.unlink()
        except Exception:
            pass
    st.cache_data.clear()


def delete_active_file():
    if ACTIVE_FILE.exists():
        ACTIVE_FILE.unlink()
    if META_FILE.exists():
        META_FILE.unlink()
    clear_cache_files()


def cache_paths():
    return {
        "op": CACHE_DIR / "op.parquet",
        "co": CACHE_DIR / "co.parquet",
        "diag": CACHE_DIR / "diag.parquet",
        "meta": CACHE_DIR / "cache_meta.json",
    }


def cache_valid():
    paths = cache_paths()
    if not ACTIVE_FILE.exists() or not paths["meta"].exists() or not paths["op"].exists() or not paths["co"].exists():
        return False
    try:
        meta = json.loads(paths["meta"].read_text(encoding="utf-8"))
        return float(meta.get("mtime", 0)) == float(ACTIVE_FILE.stat().st_mtime) and meta.get("version") == APP_CACHE_VERSION
    except Exception:
        return False




def normalize_selected_date(x):
    d = parse_date(x)
    return d.date() if pd.notna(d) else date.today()

def normalize_commercial_df(co):
    """Normaliza fecha y tienda comercial antes de usar en reportes."""
    if co is None or co.empty:
        return co
    co = co.copy()

    if "Fecha" in co.columns:
        co["Fecha"] = co["Fecha"].apply(parse_date)
        co = co[co["Fecha"].notna()]
        co["Fecha_txt"] = co["Fecha"].dt.strftime("%Y-%m-%d")

    if "Tienda" in co.columns:
        co["Tienda"] = co["Tienda"].map(canon_store)
        co = co[co["Tienda"].astype(str).str.len() > 0]

    for c in ["Dev_Pzs", "Vta_Pzs", "Vta_Imp", "Costo_Dev"]:
        if c not in co.columns:
            co[c] = 0
        co[c] = pd.to_numeric(co[c], errors="coerce").fillna(0)

    group_cols = [c for c in ["Hoja", "Fecha", "Fecha_txt", "Tienda", "ID", "Color"] if c in co.columns]
    if "Fecha" in group_cols and "Tienda" in group_cols:
        co = co.groupby(group_cols, as_index=False)[["Dev_Pzs", "Vta_Pzs", "Vta_Imp", "Costo_Dev"]].sum()
        co["Semana ISO"] = co["Fecha"].dt.isocalendar().week.astype(int)
        co["Mes"] = co["Fecha"].dt.to_period("M").astype(str)

    return co

def write_cache(op, co, diag):
    paths = cache_paths()
    op.to_parquet(paths["op"], index=False)
    co.to_parquet(paths["co"], index=False)
    diag = diag.copy()
    for _c in diag.columns:
        if diag[_c].dtype == "object":
            diag[_c] = diag[_c].astype(str)
    diag = diag.copy()
    for _c in diag.columns:
        if diag[_c].dtype == "object":
            diag[_c] = diag[_c].astype(str)
    diag.to_parquet(paths["diag"], index=False)
    paths["meta"].write_text(
        json.dumps({"mtime": ACTIVE_FILE.stat().st_mtime, "version": APP_CACHE_VERSION, "procesado": datetime.now(MX_TZ).strftime("%Y-%m-%d %H:%M:%S")}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


@st.cache_data(show_spinner=False)
def read_cache(mtime):
    paths = cache_paths()
    op = pd.read_parquet(paths["op"]) if paths["op"].exists() else pd.DataFrame()
    co = pd.read_parquet(paths["co"]) if paths["co"].exists() else pd.DataFrame()
    diag = pd.read_parquet(paths["diag"]) if paths["diag"].exists() else pd.DataFrame()
    op = normalize_operation_df(op)
    co = normalize_commercial_df(co)
    return op, co, diag


# ============================================================
# PROCESAMIENTO DEL EXCEL
# ============================================================
def find_col(cols, names):
    norm_cols = {norm_text(c): c for c in cols}
    for n in names:
        nn = norm_text(n)
        if nn in norm_cols:
            return norm_cols[nn]
    for c in cols:
        cn = norm_text(c)
        if any(norm_text(n) in cn for n in names):
            return c
    return None


def apply_nombre_map(op, plantilla):
    if op.empty or plantilla.empty or "Nombre" not in op.columns:
        return op
    p = plantilla.copy()
    c_nom = find_col(p.columns, ["Nombre"])
    if not c_nom:
        return op
    aliases = {}
    for full in p[c_nom].dropna().astype(str):
        first = full.split()[0]
        aliases[norm_text(first)] = full.strip()
        if norm_text(first).startswith("ELO"):
            aliases["ELO"] = full.strip()
        if norm_text(first).startswith("IVON") or norm_text(first).startswith("IVONNE"):
            aliases["IVON"] = full.strip()
    op["Nombre Real"] = op["Nombre"].astype(str).map(lambda x: aliases.get(norm_text(x), str(x).strip()))
    return op


def read_operation_sheet(file_path):
    """Lee y une las dos hojas operativas sin eliminar el histórico.

    Fuentes:
    - Resultados productividad
    - Resultados productividad 2

    La segunda hoja reconoce:
    Occurrence, Fecha, Ubicación, Tabla, Nómina, Actividad Realizada,
    Ingreso al area de acondicionado, Número de piezas, Hora Inicio y Hora Fin.
    """
    try:
        xls = pd.ExcelFile(file_path, engine="openpyxl")
        sheet_names = list(xls.sheet_names)
    except Exception as exc:
        return pd.DataFrame(), pd.DataFrame([{
            "Hoja": "Libro",
            "Tipo": "Error",
            "Estado": f"No fue posible abrir el archivo: {exc}",
        }])

    normalized = {norm_text(s): s for s in sheet_names}
    sources = []

    for wanted, tipo in [
        ("RESULTADOS PRODUCTIVIDAD", "Histórica"),
        ("RESULTADOS PRODUCTIVIDAD 2", "Nueva"),
    ]:
        real = normalized.get(wanted)
        if real:
            sources.append((tipo, real))

    # Alias permitido para pruebas o archivos previos.
    if not any(tipo == "Nueva" for tipo, _ in sources):
        alias = normalized.get("RESULTADOS POR CHECKLIST")
        if alias:
            sources.append(("Nueva", alias))

    if not sources:
        return pd.DataFrame(), pd.DataFrame([{
            "Hoja": "Resultados productividad",
            "Tipo": "Operación",
            "Estado": "No se encontraron las hojas operativas",
        }])

    frames = []
    diag = []

    for tipo, sheet in sources:
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet,
                engine="openpyxl",
                header=0,
            )
        except Exception as exc:
            diag.append({
                "Hoja": sheet,
                "Tipo": tipo,
                "Estado": f"Error de lectura: {exc}",
            })
            continue

        # Limpiar encabezados invisibles o espacios.
        df.columns = [str(c).strip() for c in df.columns]

        c_occ = find_col(df.columns, ["Occurrence", "Ocurrence", "Ocurrencia", "Folio"])
        c_fecha = find_col(df.columns, ["Fecha", "Fecha s", "Fecha captura"])
        c_tienda = find_col(df.columns, ["Tienda", "Ubicación", "Ubicacion", "Sucursal"])
        c_tabla = find_col(df.columns, ["Tabla"])
        c_nombre = find_col(df.columns, ["Nombre", "Nómina", "Nomina", "Colaborador", "Usuario"])
        c_actividad = find_col(df.columns, ["Actividad Realizada", "Actividad"])
        c_motivo = find_col(df.columns, [
            "Motivo de ingreso",
            "Ingreso al area de acondicionado",
            "Ingreso al área de acondicionado",
            "Motivo",
        ])
        c_piezas = find_col(df.columns, [
            "Número de piezas", "Numero de piezas",
            "Número de Piezas", "Numero de Piezas",
            "Piezas", "Cantidad",
        ])

        missing = []
        for label, col in [
            ("Fecha", c_fecha),
            ("Tienda/Ubicación", c_tienda),
            ("Actividad", c_actividad),
            ("Motivo", c_motivo),
            ("Número de piezas", c_piezas),
        ]:
            if col is None:
                missing.append(label)

        if missing:
            diag.append({
                "Hoja": sheet,
                "Tipo": tipo,
                "Estado": "Faltan columnas: " + ", ".join(missing),
                "Encabezados encontrados": " | ".join(df.columns.astype(str).tolist()),
            })
            continue

        op = pd.DataFrame({
            "Occurrence": df[c_occ].astype(str).str.strip() if c_occ else "",
            "Fecha": df[c_fecha].map(parse_date),
            "Tienda": df[c_tienda].map(canon_store),
            "Tabla": df[c_tabla].astype(str).str.strip() if c_tabla else "",
            "Nombre": df[c_nombre].astype(str).str.strip() if c_nombre else "",
            "Actividad": df[c_actividad].astype(str).str.strip(),
            "Motivo": df[c_motivo].astype(str).str.strip(),
            "Piezas": df[c_piezas].map(safe_num),
            "Hoja origen": sheet,
            "Prioridad fuente": 2 if tipo == "Nueva" else 1,
        })

        op = op.dropna(subset=["Fecha"])
        op = op[op["Tienda"].astype(str).str.strip().ne("")]
        op = op[op["Actividad"].map(norm_text).ne("")]
        op = op[pd.to_numeric(op["Piezas"], errors="coerce").fillna(0).ge(0)]

        # No recortar la hoja histórica: se conserva todo lo anterior.
        # Tampoco se recorta la hoja nueva; la deduplicación decide qué registro conservar.
        op["Semana ISO"] = op["Fecha"].dt.isocalendar().week.astype(int)
        op["Año ISO"] = op["Fecha"].dt.isocalendar().year.astype(int)
        op["Mes"] = op["Fecha"].dt.to_period("M").astype(str)

        frames.append(op)
        diag.append({
            "Hoja": sheet,
            "Tipo": tipo,
            "Estado": "OK",
            "Filas leídas": len(df),
            "Filas válidas": len(op),
            "Fecha mínima": op["Fecha"].min().strftime("%Y-%m-%d") if not op.empty else "",
            "Fecha máxima": op["Fecha"].max().strftime("%Y-%m-%d") if not op.empty else "",
            "Actividad": c_actividad,
            "Motivo": c_motivo,
            "Piezas": c_piezas,
            "Tienda": c_tienda,
        })

    if not frames:
        return pd.DataFrame(), pd.DataFrame(diag)

    result = pd.concat(frames, ignore_index=True)

    # La nueva hoja tiene prioridad únicamente cuando el registro realmente se repite.
    result = result.sort_values("Prioridad fuente")
    dedupe_cols = [
        "Occurrence", "Fecha", "Tienda", "Actividad", "Motivo", "Piezas", "Nombre"
    ]
    dedupe_cols = [c for c in dedupe_cols if c in result.columns]
    result = result.drop_duplicates(subset=dedupe_cols, keep="last")
    result = result.drop(columns=["Prioridad fuente"], errors="ignore")

    result = normalize_operation_df(result)

    # Resumen total para que el diagnóstico confirme la unión.
    diag.insert(0, {
        "Hoja": "TOTAL OPERACIÓN",
        "Tipo": "Consolidado",
        "Estado": "OK",
        "Filas leídas": "",
        "Filas válidas": len(result),
        "Fecha mínima": result["Fecha"].min().strftime("%Y-%m-%d") if not result.empty else "",
        "Fecha máxima": result["Fecha"].max().strftime("%Y-%m-%d") if not result.empty else "",
        "Actividad": "",
        "Motivo": "",
        "Piezas": f"{pd.to_numeric(result['Piezas'], errors='coerce').fillna(0).sum():,.0f}",
        "Tienda": f"{result['Tienda'].nunique()} tiendas",
    })

    return result, pd.DataFrame(diag)


def read_plantilla(file_path):
    try:
        return pd.read_excel(file_path, sheet_name="Plantilla", engine="openpyxl")
    except Exception:
        return pd.DataFrame()


def read_monthly_dev(file_path, progress=None):
    """Lector comercial por bloques con fecha normalizada y tienda homologada antes de agrupar."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    monthly_sheets = [
        s for s in wb.sheetnames
        if norm_text(s) not in ["RESULTADOS PRODUCTIVIDAD", "RESULTADOS PRODUCTIVIDAD 2", "RESULTADOS POR CHECKLIST", "PLANTILLA"]
        and re.search(r"(ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPT|OCT|NOV|DIC|ENERO|FEBR|MARZO|26|25)", norm_text(s))
    ]

    all_records = []
    diag_rows = []
    sample_rows = []
    total_sheets = max(1, len(monthly_sheets))

    for idx_sheet, sheet_name in enumerate(monthly_sheets, start=1):
        if progress:
            progress.progress(min(idx_sheet / total_sheets, 0.95), text=f"Leyendo comercial: {sheet_name}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            diag_rows.append({"Tipo": "Resumen", "Hoja": sheet_name, "Estado": "Hoja sin datos", "Registros": 0, "Dev Pzs": 0})
            continue

        max_cols = max(len(r) for r in rows[:30])
        top_rows = [list(r) + [None] * (max_cols - len(r)) for r in rows[:30]]

        header_idx = None
        tienda_col = None
        for ridx, row in enumerate(top_rows):
            tienda_cols = [i for i, v in enumerate(row) if norm_text(v) in ["TIENDA", "TIENDAS"]]
            has_dev = any(("DEV" in norm_text(v) and "PZS" in norm_text(v)) for v in row)
            if tienda_cols and has_dev:
                header_idx = ridx
                tienda_col = tienda_cols[0]
                break

        if header_idx is None or tienda_col is None:
            diag_rows.append({"Tipo": "Resumen", "Hoja": sheet_name, "Estado": "No encontró Tienda/Tiendas + Dev Pzs", "Registros": 0, "Dev Pzs": 0})
            continue

        header_row = list(rows[header_idx]) + [None] * (max_cols - len(rows[header_idx]))
        date_row = list(rows[header_idx - 1]) + [None] * (max_cols - len(rows[header_idx - 1])) if header_idx > 0 else [None] * max_cols

        date_by_col = {}
        current_date = pd.NaT
        for c in range(max_cols):
            d = parse_date(date_row[c])
            if pd.notna(d):
                current_date = d
            date_by_col[c] = current_date

        blocks = {}
        for c, h in enumerate(header_row):
            hnorm = norm_text(h)
            fecha = date_by_col.get(c, pd.NaT)
            if pd.isna(fecha):
                continue
            fecha = pd.to_datetime(fecha).normalize()

            if "DEV" in hnorm and "PZS" in hnorm:
                blocks.setdefault(fecha, {})["dev_col"] = c
            elif ("VENTA" in hnorm or "VENTAS" in hnorm) and ("PZS" in hnorm or "NETA" in hnorm) and "$" not in str(h):
                blocks.setdefault(fecha, {})["vta_pzs_col"] = c
            elif ("VENTA" in hnorm or "NETA" in hnorm) and ("$" in str(h) or "IMP" in hnorm or " EN " in f" {hnorm} "):
                blocks.setdefault(fecha, {})["vta_imp_col"] = c

        blocks = {fecha: cols for fecha, cols in blocks.items() if cols}

        if not blocks:
            diag_rows.append({
                "Tipo": "Resumen", "Hoja": sheet_name, "Estado": "No encontró bloques comerciales",
                "Fila encabezado": header_idx + 1, "Col Tienda": excel_col_name(tienda_col) if "excel_col_name" in globals() else tienda_col + 1,
                "Registros": 0, "Dev Pzs": 0
            })
            continue

        for fecha, cols in sorted(blocks.items(), key=lambda x: x[0]):
            diag_rows.append({
                "Tipo": "Columnas detectadas",
                "Hoja": sheet_name,
                "Fecha": pd.to_datetime(fecha).strftime("%Y-%m-%d"),
                "Fila encabezado": header_idx + 1,
                "Col Tienda": excel_col_name(tienda_col) if "excel_col_name" in globals() else tienda_col + 1,
                "Col Ventas Pzs": excel_col_name(cols["vta_pzs_col"]) if "vta_pzs_col" in cols and "excel_col_name" in globals() else cols.get("vta_pzs_col", ""),
                "Col Dev Pzs": excel_col_name(cols["dev_col"]) if "dev_col" in cols and "excel_col_name" in globals() else cols.get("dev_col", ""),
                "Col Venta $": excel_col_name(cols["vta_imp_col"]) if "vta_imp_col" in cols and "excel_col_name" in globals() else cols.get("vta_imp_col", ""),
            })

        acc = {}
        sheet_dev = 0.0
        sheet_vta_pzs = 0.0
        sheet_vta_imp = 0.0
        lecturas = 0
        tiendas = set()
        samples_per_sheet = 0

        for excel_row_num, raw in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            row = list(raw) + [None] * (max_cols - len(raw))
            if tienda_col >= len(row):
                continue

            raw_tienda = row[tienda_col]
            tienda = canon_store(raw_tienda)
            if not tienda:
                continue
            tiendas.add(tienda)

            for fecha, cols in blocks.items():
                fecha_norm = pd.to_datetime(fecha).normalize()

                dev_raw = row[cols["dev_col"]] if "dev_col" in cols and cols["dev_col"] < len(row) else None
                vta_raw = row[cols["vta_pzs_col"]] if "vta_pzs_col" in cols and cols["vta_pzs_col"] < len(row) else None
                imp_raw = row[cols["vta_imp_col"]] if "vta_imp_col" in cols and cols["vta_imp_col"] < len(row) else None

                dev = safe_num(dev_raw)
                vta_pzs = safe_num(vta_raw)
                vta_imp = safe_num(imp_raw)

                if dev == 0 and vta_pzs == 0 and vta_imp == 0:
                    continue

                key = (sheet_name, fecha_norm, tienda)
                if key not in acc:
                    acc[key] = {"Dev_Pzs": 0.0, "Vta_Pzs": 0.0, "Vta_Imp": 0.0}
                acc[key]["Dev_Pzs"] += dev
                acc[key]["Vta_Pzs"] += vta_pzs
                acc[key]["Vta_Imp"] += vta_imp

                sheet_dev += dev
                sheet_vta_pzs += vta_pzs
                sheet_vta_imp += vta_imp
                lecturas += 1

                raw_norm = norm_text(raw_tienda)
                if samples_per_sheet < 350 and (dev != 0 or "MIRAVALLE" in raw_norm or "GUADALAJARA" in raw_norm or "ATEMAJAC" in raw_norm):
                    sample_rows.append({
                        "Hoja": sheet_name,
                        "Fila Excel": excel_row_num,
                        "Fecha": fecha_norm.strftime("%Y-%m-%d"),
                        "Tienda cruda": str(raw_tienda),
                        "Tienda homologada": tienda,
                        "Col Dev": excel_col_name(cols["dev_col"]) if "dev_col" in cols and "excel_col_name" in globals() else cols.get("dev_col", ""),
                        "Dev crudo": str(dev_raw),
                        "Dev num": dev,
                        "Ventas crudo": str(vta_raw),
                        "Ventas num": vta_pzs,
                        "Venta $ crudo": str(imp_raw),
                        "Venta $ num": vta_imp,
                    })
                    samples_per_sheet += 1

        for (hoja, fecha, tienda), vals in acc.items():
            fecha = pd.to_datetime(fecha).normalize()
            tienda = canon_store(tienda)
            all_records.append({
                "Hoja": hoja,
                "Fecha": fecha,
                "Fecha_txt": fecha.strftime("%Y-%m-%d"),
                "Tienda": tienda,
                "Dev_Pzs": vals["Dev_Pzs"],
                "Vta_Pzs": vals["Vta_Pzs"],
                "Vta_Imp": vals["Vta_Imp"],
                "Costo_Dev": 0.0,
                "ID": "",
                "Color": "",
            })

        diag_rows.append({
            "Tipo": "Resumen",
            "Hoja": sheet_name,
            "Estado": "OK",
            "Fila encabezado": header_idx + 1,
            "Fila fechas": header_idx,
            "Col Tienda": excel_col_name(tienda_col) if "excel_col_name" in globals() else tienda_col + 1,
            "Fechas detectadas": len(blocks),
            "Registros agrupados": len(acc),
            "Lecturas con valor": lecturas,
            "Tiendas detectadas": len(tiendas),
            "Dev Pzs": sheet_dev,
            "Venta Pzs": sheet_vta_pzs,
            "Venta $": sheet_vta_imp,
        })

    wb.close()

    co = pd.DataFrame(all_records)
    if not co.empty:
        co["Fecha"] = co["Fecha"].apply(parse_date)
        co["Fecha_txt"] = co["Fecha"].dt.strftime("%Y-%m-%d")
        co["Tienda"] = co["Tienda"].map(canon_store)
        # Reagrupar después de homologar para unir Guadalajara Miravalle -> Miravalle y Guadalajara -> Atemajac.
        co = co.groupby(["Hoja", "Fecha", "Fecha_txt", "Tienda", "ID", "Color"], as_index=False)[["Dev_Pzs", "Vta_Pzs", "Vta_Imp", "Costo_Dev"]].sum()
        co["Semana ISO"] = co["Fecha"].dt.isocalendar().week.astype(int)
        co["Mes"] = co["Fecha"].dt.to_period("M").astype(str)

    diag = pd.DataFrame(diag_rows)
    samples = pd.DataFrame(sample_rows)
    if not samples.empty:
        samples.insert(0, "Tipo", "Muestra lectura")
        diag = pd.concat([diag, samples], ignore_index=True, sort=False)

    co = normalize_commercial_df(co)
    return co, diag



def process_excel(file_path):
    progress = st.progress(0, text="Iniciando procesamiento...")
    progress.progress(0.10, text="Leyendo Resultados productividad y Resultados productividad 2...")
    op, diag_op = read_operation_sheet(file_path)

    progress.progress(0.25, text="Leyendo Plantilla...")
    plantilla = read_plantilla(file_path)
    op = apply_nombre_map(op, plantilla)

    progress.progress(0.35, text="Leyendo hojas mensuales Dev Pzs...")
    co, diag_co = read_monthly_dev(file_path, progress=progress)

    progress.progress(0.90, text="Guardando cache optimizado...")
    diag = pd.concat([diag_op, diag_co], ignore_index=True)

    # Validación de fechas operativas para detectar inversiones mes/día.
    if op is not None and not op.empty and "Fecha" in op.columns:
        fechas_op = pd.to_datetime(op["Fecha"], errors="coerce").dropna()
        if not fechas_op.empty:
            diag_fecha = pd.DataFrame([{
                "Hoja": "VALIDACIÓN FECHAS OPERACIÓN",
                "Tipo": "Control",
                "Estado": "OK",
                "Filas válidas": len(fechas_op),
                "Fecha mínima": fechas_op.min().strftime("%Y-%m-%d"),
                "Fecha máxima": fechas_op.max().strftime("%Y-%m-%d"),
            }])
            diag = pd.concat([diag_fecha, diag], ignore_index=True)
    op = normalize_operation_df(op)
    co = normalize_commercial_df(co)
    write_cache(op, co, diag)
    progress.progress(1.0, text="Archivo procesado correctamente.")
    op = normalize_operation_df(op)
    co = normalize_commercial_df(co)
    return op, co, diag



def split_operation(op):
    if op is None or op.empty:
        return op
    df = op.copy()
    act = df["Actividad"].map(norm_text)
    mot = df["Motivo"].map(norm_text)

    es_recoleccion_muertos = act.str.contains(
        r"RECOLECCION DE MUERTOS|RECOLECCIÓN DE MUERTOS",
        regex=True,
        na=False,
    )
    es_motivo_muertos = mot.str.contains("MUERTO", na=False)

    # Muertos: únicamente Recolección de muertos + motivo Muertos.
    df["Muertos"] = np.where(
        es_recoleccion_muertos & es_motivo_muertos,
        df["Piezas"],
        0,
    )

    # Cajas y Probador son ingresos, no cualquier actividad posterior.
    es_ingreso_o_recoleccion = act.str.contains(
        r"^INGRESO$|RECOLECCION|RECOLECCIÓN",
        regex=True,
        na=False,
    )
    df["Cajas"] = np.where(
        es_ingreso_o_recoleccion & mot.str.contains("CAJA", na=False),
        df["Piezas"],
        0,
    )
    df["Probador"] = np.where(
        es_ingreso_o_recoleccion
        & (
            mot.str.contains("PROBADOR", na=False)
            | act.str.contains("PROBADOR", na=False)
        ),
        df["Piezas"],
        0,
    )

    # Recolectadas considera Recolección e Ingreso de la nueva fuente.
    df["Recolectadas"] = np.where(
        act.str.contains(r"RECOLECCION|RECOLECCIÓN|^INGRESO$", regex=True, na=False),
        df["Piezas"],
        0,
    )
    df["Habilitadas"] = np.where(
        act.str.contains(r"ACONDICION|HABILIT", regex=True, na=False),
        df["Piezas"],
        0,
    )
    df["Ubicadas"] = np.where(
        act.str.contains(r"UBIC", regex=True, na=False),
        df["Piezas"],
        0,
    )
    return df


def filter_stores(df, stores=None):
    if df.empty or not stores:
        return df
    return df[df["Tienda"].isin(stores)]



def normalize_operation_df(op):
    """Normaliza fechas/tiendas operativas antes de reportar."""
    if op is None or op.empty:
        return op
    op = op.copy()

    # Recalcular fecha desde columnas originales si están disponibles.
    for cand in ["Fecha s", "Fecha_s", "Fecha original", "Fecha Original"]:
        if cand in op.columns:
            parsed = op[cand].apply(parse_date)
            if parsed.notna().sum() >= max(1, len(op) * 0.4):
                op["Fecha"] = parsed
                break

    if "Fecha" in op.columns:
        op["Fecha"] = op["Fecha"].apply(parse_date)
        op = op[op["Fecha"].notna()]
        op["Semana ISO"] = op["Fecha"].dt.isocalendar().week.astype(int)
        op["Mes"] = op["Fecha"].dt.to_period("M").astype(str)

    if "Tienda" in op.columns:
        op["Tienda"] = op["Tienda"].map(canon_store)

    return op


def filter_commercial_by_date(co, start, end, stores_list):
    if co is None or co.empty:
        return pd.DataFrame()
    co = normalize_commercial_df(co)
    co = filter_stores(co, stores_list)
    if co.empty:
        return co

    start_txt = pd.to_datetime(start).strftime("%Y-%m-%d")
    end_txt = pd.to_datetime(end).strftime("%Y-%m-%d")

    if "Fecha_txt" not in co.columns:
        co["Fecha_txt"] = pd.to_datetime(co["Fecha"], errors="coerce").dt.strftime("%Y-%m-%d")

    out = co[(co["Fecha_txt"] >= start_txt) & (co["Fecha_txt"] <= end_txt)].copy()

    # Rescate: si no encontró Dev y es un solo día, intenta fecha con día/mes invertido.
    if out.empty and start_txt == end_txt:
        try:
            d = pd.to_datetime(start)
            alt = pd.Timestamp(year=d.year, month=d.day, day=d.month)
            alt_txt = alt.strftime("%Y-%m-%d")
            out = co[co["Fecha_txt"].eq(alt_txt)].copy()
        except Exception:
            pass

    return out

def closing_pending_by_store(op, co, cutoff_date, stores=None):
    """Calcula el saldo pendiente real por tienda hasta una fecha de cierre.

    Saldo diario:
        saldo final = máximo(saldo anterior + ingresos del día - ubicadas del día, 0)

    Esto evita que las ubicaciones excedentes de una tienda compensen el
    pendiente de otra y permite trasladar correctamente el cierre del domingo
    a la semana siguiente.
    """
    op = normalize_operation_df(op)
    co = normalize_commercial_df(co)
    stores_list = stores or PROJECT_STORES
    cutoff = parse_date(cutoff_date)

    result = {store: 0.0 for store in stores_list}
    if pd.isna(cutoff):
        return result

    op2 = split_operation(op)
    min_dates = []

    if op2 is not None and not op2.empty and "Fecha" in op2.columns:
        valid_op_dates = pd.to_datetime(op2["Fecha"], errors="coerce").dropna()
        if not valid_op_dates.empty:
            min_dates.append(valid_op_dates.min().normalize())

    if co is not None and not co.empty and "Fecha" in co.columns:
        valid_co_dates = pd.to_datetime(co["Fecha"], errors="coerce").dropna()
        if not valid_co_dates.empty:
            min_dates.append(valid_co_dates.min().normalize())

    if not min_dates:
        return result

    first_date = min(min_dates)
    if cutoff < first_date:
        return result

    op_cut = (
        op2[
            (pd.to_datetime(op2["Fecha"], errors="coerce") >= first_date)
            & (pd.to_datetime(op2["Fecha"], errors="coerce") <= cutoff)
        ].copy()
        if op2 is not None and not op2.empty
        else pd.DataFrame()
    )
    op_cut = filter_stores(op_cut, stores_list)

    co_cut = filter_commercial_by_date(co, first_date, cutoff, stores_list)

    daily_parts = []

    if not op_cut.empty:
        op_cut["Fecha"] = pd.to_datetime(op_cut["Fecha"], errors="coerce").dt.normalize()
        op_daily = (
            op_cut.groupby(["Fecha", "Tienda"], as_index=False)
            .agg({
                "Muertos": "sum",
                "Cajas": "sum",
                "Probador": "sum",
                "Ubicadas": "sum",
            })
        )
        op_daily["Ingresos operación"] = (
            pd.to_numeric(op_daily["Muertos"], errors="coerce").fillna(0)
            + pd.to_numeric(op_daily["Cajas"], errors="coerce").fillna(0)
            + pd.to_numeric(op_daily["Probador"], errors="coerce").fillna(0)
        )
        daily_parts.append(
            op_daily[["Fecha", "Tienda", "Ingresos operación", "Ubicadas"]]
        )

    if not co_cut.empty:
        co_cut["Fecha"] = pd.to_datetime(co_cut["Fecha"], errors="coerce").dt.normalize()
        co_daily = (
            co_cut.groupby(["Fecha", "Tienda"], as_index=False)["Dev_Pzs"]
            .sum()
            .rename(columns={"Dev_Pzs": "Dev diario"})
        )
    else:
        co_daily = pd.DataFrame(columns=["Fecha", "Tienda", "Dev diario"])

    if daily_parts:
        daily = daily_parts[0]
    else:
        daily = pd.DataFrame(
            columns=["Fecha", "Tienda", "Ingresos operación", "Ubicadas"]
        )

    daily = daily.merge(co_daily, on=["Fecha", "Tienda"], how="outer")
    for col in ["Ingresos operación", "Ubicadas", "Dev diario"]:
        daily[col] = pd.to_numeric(daily.get(col, 0), errors="coerce").fillna(0)

    daily["Ingresos"] = daily["Ingresos operación"] + daily["Dev diario"]
    daily = daily.sort_values(["Tienda", "Fecha"])

    for store in stores_list:
        saldo = 0.0
        store_daily = daily[daily["Tienda"].eq(store)]
        for row in store_daily.itertuples(index=False):
            saldo = max(
                saldo + float(row.Ingresos) - float(row.Ubicadas),
                0.0,
            )
        result[store] = saldo

    return result

def table_by_store(
    op,
    co,
    start_date,
    end_date,
    stores=None,
    carryover_mode="previous_day",
):
    """Construye la tabla por tienda para un periodo.

    carryover_mode:
    - "previous_day": traslada el saldo acumulado al cierre del día anterior.
    - "previous_sunday": traslada el saldo acumulado al domingo anterior.
    - "none": no agrega saldo anterior; se usan solo ingresos del periodo.
    """
    op = normalize_operation_df(op)
    co = normalize_commercial_df(co)

    op2 = split_operation(op)
    start = parse_date(start_date)
    end = parse_date(end_date)
    stores_list = stores or PROJECT_STORES

    op_p = (
        op2[(op2["Fecha"] >= start) & (op2["Fecha"] <= end)]
        if op2 is not None and not op2.empty
        else pd.DataFrame()
    )
    op_p = filter_stores(op_p, stores_list)
    co_p = filter_commercial_by_date(co, start, end, stores_list)

    if carryover_mode == "none":
        prior_balances = {store: 0.0 for store in stores_list}
    else:
        # Para una semana ISO el día anterior al lunes es exactamente el domingo previo.
        cutoff = start - pd.Timedelta(days=1)
        prior_balances = closing_pending_by_store(
            op,
            co,
            cutoff,
            stores_list,
        )

    rows = []
    for t in stores_list:
        dev = (
            pd.to_numeric(
                co_p.loc[co_p["Tienda"].eq(t), "Dev_Pzs"],
                errors="coerce",
            ).fillna(0).sum()
            if not co_p.empty and "Dev_Pzs" in co_p.columns
            else 0
        )

        o = op_p[op_p["Tienda"].eq(t)] if not op_p.empty else pd.DataFrame()

        muertos = o["Muertos"].sum() if not o.empty and "Muertos" in o.columns else 0
        cajas = o["Cajas"].sum() if not o.empty and "Cajas" in o.columns else 0
        prob = o["Probador"].sum() if not o.empty and "Probador" in o.columns else 0
        reco = o["Recolectadas"].sum() if not o.empty and "Recolectadas" in o.columns else 0
        hab = o["Habilitadas"].sum() if not o.empty and "Habilitadas" in o.columns else 0
        ubic = o["Ubicadas"].sum() if not o.empty and "Ubicadas" in o.columns else 0

        ingresos_periodo = dev + muertos + cajas + prob
        pend_ant = float(prior_balances.get(t, 0))
        total_base = ingresos_periodo + pend_ant

        pend_hab = max(total_base - hab, 0)
        pend_ub = max(total_base - ubic, 0)
        procesado = max(total_base - pend_ub, 0)

        pct_hab = min(hab / total_base * 100, 100) if total_base else 0
        pct_ub = min(procesado / total_base * 100, 100) if total_base else 0

        rows.append({
            "Tienda": t,
            "Dev pzs": dev,
            "Muertos": muertos,
            "Cajas": cajas,
            "Probador": prob,
            "Ingresos periodo": ingresos_periodo,
            "Pend. Ant.": pend_ant,
            "Total": total_base,
            "Recolectadas": reco,
            "Habilitadas": hab,
            "Pend. Hab.": pend_hab,
            "% Acond.": pct_hab,
            "Ubicadas": ubic,
            "Pend. Ub.": pend_ub,
            "% Ubic.": pct_ub,
        })

    return pd.DataFrame(rows)


def summary_from_table(df, income_column="Total"):
    """Calcula KPI respetando los pendientes individuales por tienda.

    - Pendiente general = suma de `Pend. Ub.` de cada tienda.
    - % Procesado = (base - pendiente) / base.
    - Nunca se compensan pendientes entre tiendas.
    """
    if df is None or df.empty:
        return {
            "Ingresos": 0,
            "Acondicionado": 0,
            "Ubicado": 0,
            "Pendiente": 0,
            "% Procesado": 0,
        }

    base_col = income_column if income_column in df.columns else "Total"

    ingresos = pd.to_numeric(
        df.get(base_col, pd.Series(dtype=float)),
        errors="coerce",
    ).fillna(0).sum()

    hab = pd.to_numeric(
        df.get("Habilitadas", pd.Series(dtype=float)),
        errors="coerce",
    ).fillna(0).sum()

    ubic = pd.to_numeric(
        df.get("Ubicadas", pd.Series(dtype=float)),
        errors="coerce",
    ).fillna(0).sum()

    pendiente = pd.to_numeric(
        df.get("Pend. Ub.", pd.Series(dtype=float)),
        errors="coerce",
    ).fillna(0).clip(lower=0).sum()

    procesado = max(float(ingresos) - float(pendiente), 0)
    pct = min(procesado / float(ingresos) * 100, 100) if ingresos > 0 else 0

    return {
        "Ingresos": ingresos,
        "Acondicionado": hab,
        "Ubicado": ubic,
        "Pendiente": pendiente,
        "% Procesado": pct,
    }


def format_display(df):
    if df is None or df.empty:
        return df
    out = df.copy()
    for c in out.columns:
        if c == "Tienda" or c in ["Nombre", "Nombre Real", "Actividad"]:
            continue
        if "%" in str(c):
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).map(fmt_pct)
        elif "$" in str(c) or "Importe" in str(c) or "Venta" in str(c):
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).map(fmt_money)
        else:
            ser = pd.to_numeric(out[c], errors="coerce")
            if ser.notna().mean() > 0.75:
                out[c] = ser.fillna(0).map(fmt_num)
    return out


# ============================================================
# COMPONENTES VISUALES
# ============================================================
def aggrid_table(df, height=360, editable=False, key=None):
    if df is None or df.empty:
        st.info("Sin información para mostrar.")
        return df
    show = format_display(df)
    auto_height = min(max(118 + len(show) * 34, 170), height)

    if not AGGRID_OK:
        st.dataframe(show, hide_index=True, width="stretch", height=auto_height)
        return df

    gb = GridOptionsBuilder.from_dataframe(show)
    gb.configure_default_column(filter=True, sortable=True, resizable=True, editable=editable, minWidth=105)
    if "Tienda" in show.columns:
        gb.configure_column("Tienda", pinned="left", minWidth=145)
    for col in show.columns:
        if col != "Tienda":
            if col == "% Ubic.":
                gb.configure_column(
                    col,
                    type=["rightAligned"],
                    minWidth=105,
                    cellStyle=JsCode("""
                        function(params) {
                            const v = parseFloat(String(params.value).replace('%','').replace(',',''));
                            if (isNaN(v)) return {};
                            if (v < 75) {
                                return {'color':'#D71920','fontWeight':'900'};
                            }
                            if (v >= 90) {
                                return {'color':'#008A3B','fontWeight':'900'};
                            }
                            return {'color':'#111827','fontWeight':'700'};
                        }
                    """)
                )
            else:
                gb.configure_column(col, type=["rightAligned"], minWidth=105)
    opts = gb.build()
    opts["rowHeight"] = 34
    opts["headerHeight"] = 38
    opts["enableCellTextSelection"] = True
    opts["suppressRowClickSelection"] = True
    opts["getRowStyle"] = JsCode("""
        function(params) {
            if (params.node.rowIndex % 2 === 0) { return {'backgroundColor':'#FFFFFF'}; }
            return {'backgroundColor':'#F8FAFC'};
        }
    """)
    css = {
        ".ag-header": {"background-color": f"{AZUL} !important"},
        ".ag-header-cell": {"background-color": f"{AZUL} !important", "color": "#FFFFFF !important", "font-weight": "900 !important"},
        ".ag-header-cell-text": {"color": "#FFFFFF !important", "font-weight": "900 !important"},
        ".ag-icon": {"color": "#FFFFFF !important", "fill": "#FFFFFF !important"},
        ".ag-root-wrapper": {"border": "1px solid #E1E7F0 !important", "border-radius": "10px !important", "overflow": "hidden !important"},
        ".ag-cell": {"font-size": "12px !important"},
    }
    result = AgGrid(
        show,
        gridOptions=opts,
        height=auto_height,
        width="100%",
        fit_columns_on_grid_load=True,
        allow_unsafe_jscode=True,
        custom_css=css,
        theme="alpine",
        key=key or f"ag_{abs(hash(str(show.columns.tolist())+str(len(show))))}",
            )
    if editable and result and "data" in result:
        return pd.DataFrame(result["data"])
    return df


def panel(title, df, height=360, editable=False):
    st.markdown(f'<div class="panel-title">{title}</div>', unsafe_allow_html=True)
    return aggrid_table(df, height=height, editable=editable, key=f"panel_{norm_text(title)}")


def kpis(res):
    vals = [
        ("↻", "Piezas Ingresadas", fmt_num(res.get("Ingresos", 0)), "Dev + muertos + cajas + probador", ROSA),
        ("✓", "Piezas Acondicionadas", fmt_num(res.get("Acondicionado", 0)), "Acondicionado", "#3720B8"),
        ("⊕", "Piezas Ubicadas", fmt_num(res.get("Ubicado", 0)), "Ubicado", "#F59E0B"),
        ("⌛", "Pendientes por Ubicar", fmt_num(res.get("Pendiente", 0)), "Piezas ingresadas - piezas ubicadas", "#05B957"),
        ("%", "% Procesado", fmt_pct(res.get("% Procesado", 0)), "Piezas ubicadas / piezas ingresadas", "#3720B8"),
    ]
    html = '<div class="ps-kpi-grid">'
    for icon, title, val, sub, color in vals:
        html += (
            '<div class="ps-kpi-card">'
            f'<div class="ps-kpi-icon" style="background:{color};">{icon}</div>'
            '<div><div class="ps-kpi-title">'+title+'</div>'
            f'<div class="ps-kpi-value">{val}</div>'
            f'<div class="ps-kpi-sub">{sub}</div></div></div>'
        )
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


def combined_chart(df, title, income_column="Total"):
    if df is None or df.empty:
        return

    chart_df = df.copy()
    for c in [income_column, "Habilitadas", "Ubicadas"]:
        if c not in chart_df.columns:
            chart_df[c] = 0
        chart_df[c] = pd.to_numeric(chart_df[c], errors="coerce").fillna(0)

    raw_max = max(float(chart_df[c].max()) for c in [income_column, "Habilitadas", "Ubicadas"])
    ymax = raw_max * 1.55 if raw_max > 0 else 10
    leader_gap = max(ymax * 0.075, 30)

    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=chart_df["Tienda"],
        y=chart_df[income_column],
        mode="lines+markers",
        name="Total ingresos",
        line=dict(color="#43A5FF", width=4),
        marker=dict(color="#43A5FF", size=9),
        hovertemplate="<b>%{x}</b><br>Total ingresos: %{y:,.0f}<extra></extra>",
    ))

    fig.add_trace(go.Bar(
        x=chart_df["Tienda"],
        y=chart_df["Habilitadas"],
        name="Pzas Habilitadas",
        marker_color=AZUL,
        text=chart_df["Habilitadas"].map(lambda x: f"<b>{x:,.0f}</b>"),
        textposition="outside",
        textfont=dict(color="#111827", size=13, family="Arial Black"),
        cliponaxis=False,
        hovertemplate="<b>%{x}</b><br>Habilitadas: %{y:,.0f}<extra></extra>",
    ))

    fig.add_trace(go.Bar(
        x=chart_df["Tienda"],
        y=chart_df["Ubicadas"],
        name="Pzas Ubicadas",
        marker_color=ROSA,
        text=chart_df["Ubicadas"].map(lambda x: f"<b>{x:,.0f}</b>"),
        textposition="outside",
        textfont=dict(color="#111827", size=13, family="Arial Black"),
        cliponaxis=False,
        hovertemplate="<b>%{x}</b><br>Ubicadas: %{y:,.0f}<extra></extra>",
    ))

    for tienda, total, habilitadas, ubicadas in zip(
        chart_df["Tienda"],
        chart_df[income_column],
        chart_df["Habilitadas"],
        chart_df["Ubicadas"],
    ):
        group_top = max(float(total), float(habilitadas), float(ubicadas))
        label_y = min(group_top + leader_gap, ymax * 0.94)

        fig.add_shape(
            type="line",
            x0=tienda,
            x1=tienda,
            y0=float(total) + max(ymax * 0.012, 5),
            y1=label_y - max(ymax * 0.018, 8),
            line=dict(color="#43A5FF", width=2, dash="dot"),
            layer="above",
        )
        fig.add_annotation(
            x=tienda,
            y=label_y,
            text=f"<b>{float(total):,.0f}</b>",
            showarrow=False,
            font=dict(color="#111827", size=13, family="Arial Black"),
            bgcolor="rgba(255,255,255,0.96)",
            bordercolor="#D9E1EE",
            borderwidth=1,
            borderpad=3,
        )

    fig.update_layout(
        title=title,
        barmode="group",
        height=580,
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend=dict(orientation="h", y=1.08, x=1, xanchor="right"),
        margin=dict(l=8, r=8, t=72, b=92),
        dragmode=False,
        uniformtext_minsize=10,
        uniformtext_mode="show",
    )
    fig.update_xaxes(tickangle=-45, showgrid=False, fixedrange=True)
    fig.update_yaxes(
        showgrid=True,
        gridcolor="#E5E7EB",
        fixedrange=True,
        range=[0, ymax],
        tickformat=",d",
    )
    st.plotly_chart(
        fig,
        width="stretch",
        config={
            "displayModeBar": False,
            "scrollZoom": False,
            "doubleClick": False,
            "responsive": True,
        },
    )


def _pdf_icon(symbol, color_hex):
    d = Drawing(34, 34)
    d.add(Circle(17, 17, 16, fillColor=colors.HexColor(color_hex), strokeColor=None))
    d.add(String(
        17, 12, symbol,
        textAnchor="middle",
        fontName="Helvetica-Bold",
        fontSize=16,
        fillColor=colors.white,
    ))
    return d


def _pdf_kpi_card(symbol, title, value, note, color_hex, styles):
    icon = _pdf_icon(symbol, color_hex)

    title_p = Paragraph(
        f"<b>{title}</b>",
        ParagraphStyle(
            f"kpi_title_{re.sub(r'[^A-Za-z0-9]', '', title)}",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7.1,
            leading=8.2,
            textColor=colors.HexColor("#15102E"),
            spaceAfter=0,
        ),
    )
    value_p = Paragraph(
        f"<font color='{ROSA}' size='14'><b>{value}</b></font>",
        ParagraphStyle(
            f"kpi_value_{re.sub(r'[^A-Za-z0-9]', '', title)}",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=15,
            textColor=colors.HexColor(ROSA),
            spaceAfter=0,
        ),
    )
    note_p = Paragraph(
        note,
        ParagraphStyle(
            f"kpi_note_{re.sub(r'[^A-Za-z0-9]', '', title)}",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=5.2,
            leading=6.2,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=0,
        ),
    )

    text_stack = Table(
        [[title_p], [value_p], [note_p]],
        colWidths=[102],
        rowHeights=[17, 18, 18],
    )
    text_stack.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 1),
        ("RIGHTPADDING", (0,0), (-1,-1), 1),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))

    inner = Table(
        [[icon, text_stack]],
        colWidths=[36, 104],
        rowHeights=[58],
    )
    inner.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (0,0), 1),
        ("RIGHTPADDING", (0,0), (0,0), 3),
        ("LEFTPADDING", (1,0), (1,0), 1),
        ("RIGHTPADDING", (1,0), (1,0), 1),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))

    outer = Table([[inner]], colWidths=[146], rowHeights=[66])
    outer.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.white),
        ("BOX", (0,0), (-1,-1), 0.55, colors.HexColor("#D9E1EE")),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    return outer



def _pdf_footer(canvas, doc):
    canvas.saveState()
    page_width, _ = landscape(letter)

    canvas.setStrokeColor(colors.HexColor("#D9E1EE"))
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, 18, page_width - doc.rightMargin, 18)

    canvas.setFont("Helvetica-Bold", 6.2)
    canvas.setFillColor(colors.HexColor("#5B6476"))
    canvas.drawString(
        doc.leftMargin,
        8,
        "INFORMACIÓN CONFIDENCIAL | Price Shoes | Operaciones Ropa",
    )
    canvas.restoreState()


def _pdf_chart(df):
    drawing = Drawing(742, 215)
    x0, y0 = 48, 34
    width, height = 655, 134

    tiendas = list(df["Tienda"].astype(str))
    hab = pd.to_numeric(df["Habilitadas"], errors="coerce").fillna(0).astype(float).tolist()
    ubic = pd.to_numeric(df["Ubicadas"], errors="coerce").fillna(0).astype(float).tolist()
    total = pd.to_numeric(df["Total"], errors="coerce").fillna(0).astype(float).tolist()

    maxv = max(hab + ubic + total + [10.0])
    ymax = maxv * 1.62

    # Ejes y retícula.
    drawing.add(Line(x0, y0, x0 + width, y0, strokeColor=colors.HexColor("#AEB8C8"), strokeWidth=0.7))
    for i in range(6):
        val = ymax * i / 5
        y = y0 + height * i / 5
        drawing.add(Line(x0, y, x0 + width, y, strokeColor=colors.HexColor("#E6EAF0"), strokeWidth=0.5))
        drawing.add(String(
            x0 - 8, y - 2, f"{val:,.0f}",
            textAnchor="end", fontSize=5.5,
            fillColor=colors.HexColor("#586174"),
        ))

    n = max(1, len(tiendas))
    group_w = width / n
    bar_w = min(24, group_w * 0.27)
    line_points = []

    def add_label_box(cx, cy, text, font_size=6.5, pad_x=3.5, pad_y=2.0):
        label_w = max(18, len(text) * font_size * 0.56 + pad_x * 2)
        label_h = font_size + pad_y * 2 + 1
        drawing.add(Rect(
            cx - label_w / 2,
            cy - pad_y - 1,
            label_w,
            label_h,
            fillColor=colors.white,
            strokeColor=colors.HexColor("#D9E1EE"),
            strokeWidth=0.35,
        ))
        drawing.add(String(
            cx,
            cy,
            text,
            textAnchor="middle",
            fontName="Helvetica-Bold",
            fontSize=font_size,
            fillColor=colors.black,
        ))

    for i, tienda in enumerate(tiendas):
        center = x0 + group_w * (i + 0.5)
        x_h = center - bar_w - 1.5
        x_u = center + 1.5

        h_h = height * hab[i] / ymax
        h_u = height * ubic[i] / ymax
        point_y = y0 + height * total[i] / ymax

        drawing.add(Rect(
            x_h, y0, bar_w, h_h,
            fillColor=colors.HexColor(AZUL),
            strokeColor=None,
        ))
        drawing.add(Rect(
            x_u, y0, bar_w, h_u,
            fillColor=colors.HexColor(ROSA),
            strokeColor=None,
        ))

        # Etiquetas de barras: siempre visibles sobre fondo blanco.
        hab_label_y = y0 + h_h + 7
        ubic_label_y = y0 + h_u + 7
        add_label_box(x_h + bar_w / 2, hab_label_y, f"{hab[i]:,.0f}", font_size=6.4)
        add_label_box(x_u + bar_w / 2, ubic_label_y, f"{ubic[i]:,.0f}", font_size=6.4)

        line_points.append((center, point_y))

        # El número de la línea se coloca por encima de la barra más alta del grupo.
        group_top_value = max(total[i], hab[i], ubic[i])
        group_top_y = y0 + height * group_top_value / ymax
        label_y = min(group_top_y + 26, y0 + height + 23)

        leader = Line(
            center,
            point_y + 3,
            center,
            label_y - 8,
            strokeColor=colors.HexColor("#43A5FF"),
            strokeWidth=1.1,
        )
        leader.strokeDashArray = [2, 2]
        drawing.add(leader)
        add_label_box(center, label_y, f"{total[i]:,.0f}", font_size=6.7)

        drawing.add(String(
            center + 2,
            y0 - 13,
            tienda,
            textAnchor="end",
            fontSize=5.7,
            fillColor=colors.HexColor("#4B5563"),
            angle=35,
        ))

    if len(line_points) >= 2:
        drawing.add(PolyLine(
            line_points,
            strokeColor=colors.HexColor("#43A5FF"),
            strokeWidth=2.2,
        ))
    for x, y in line_points:
        drawing.add(Circle(
            x, y, 2.5,
            fillColor=colors.HexColor("#43A5FF"),
            strokeColor=None,
        ))

    # Leyenda.
    legend_y = 201
    drawing.add(Line(500, legend_y, 518, legend_y, strokeColor=colors.HexColor("#43A5FF"), strokeWidth=2.2))
    drawing.add(String(522, legend_y - 2, "Total ingresos", fontSize=5.8, fillColor=colors.HexColor("#313847")))
    drawing.add(Rect(588, legend_y - 4, 8, 8, fillColor=colors.HexColor(AZUL), strokeColor=None))
    drawing.add(String(600, legend_y - 2, "Pzas Habilitadas", fontSize=5.8, fillColor=colors.HexColor("#313847")))
    drawing.add(Rect(675, legend_y - 4, 8, 8, fillColor=colors.HexColor(ROSA), strokeColor=None))
    drawing.add(String(687, legend_y - 2, "Pzas Ubicadas", fontSize=5.8, fillColor=colors.HexColor("#313847")))

    return drawing


def build_pdf_report(title, subtitle, kpi_values, df):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=18,
        leftMargin=18,
        topMargin=14,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()
    story = []

    # Encabezado con logo Price Shoes.
    logo_path = ASSETS_DIR / "price_shoes_logo.png"
    logo = RLImage(str(logo_path), width=58, height=34) if logo_path.exists() else Paragraph("<b>Price Shoes</b>", styles["Normal"])
    title_block = Paragraph(
        f"<font name='Helvetica-Bold' color='#1D1259' size='13'>Indicadores Cambios y Muertos</font><br/>"
        f"<font name='Helvetica-Bold' color='#1D1259' size='10'>{title}</font>"
        f"<font name='Helvetica' color='#5B6476' size='8'> | {subtitle}</font>",
        ParagraphStyle("pdf_header", parent=styles["Normal"], leading=14),
    )
    header = Table([[logo, title_block]], colWidths=[72, 650], rowHeights=[40])
    header.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    story.append(header)

    pink_line = Table([[""]], colWidths=[744], rowHeights=[3])
    pink_line.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.HexColor(ROSA))]))
    story.append(pink_line)
    story.append(Spacer(1, 7))

    cards = [
        _pdf_kpi_card("↻", "Piezas Ingresadas", fmt_num(kpi_values.get("Ingresos", 0)), "Dev + muertos + cajas + probador", ROSA, styles),
        _pdf_kpi_card("✓", "Piezas Acondicionadas", fmt_num(kpi_values.get("Acondicionado", 0)), "Acondicionado", "#5B00D6", styles),
        _pdf_kpi_card("⊕", "Piezas Ubicadas", fmt_num(kpi_values.get("Ubicado", 0)), "Ubicado", "#F59E0B", styles),
        _pdf_kpi_card("⌛", "Pendientes por Ubicar", fmt_num(kpi_values.get("Pendiente", 0)), "Piezas ingresadas - piezas ubicadas", "#05B957", styles),
        _pdf_kpi_card("%", "% Procesado", fmt_pct(kpi_values.get("% Procesado", 0)), "Piezas ubicadas / piezas ingresadas", "#5B00D6", styles),
    ]
    cards_row = Table([cards], colWidths=[148,148,148,148,148], rowHeights=[68])
    cards_row.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 2),
        ("RIGHTPADDING", (0,0), (-1,-1), 2),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))
    story.append(cards_row)
    story.append(Spacer(1, 6))

    story.append(Paragraph(
        "<b>Tabla por tienda - Por Día</b>",
        ParagraphStyle("pdf_h2", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#1D1259"), spaceAfter=5),
    ))

    columns = ["Tienda","Dev pzs","Muertos","Cajas","Probador","Pend. Ant.","Total","Recolectadas","Habilitadas","Pend. Hab.","% Acond.","Ubicadas","Pend. Ub.","% Ubic."]
    pdf_df = df[[c for c in columns if c in df.columns]].copy()

    raw_pct_ubic = pd.to_numeric(pdf_df["% Ubic."], errors="coerce").fillna(0).tolist() if "% Ubic." in pdf_df.columns else []
    for col in pdf_df.columns:
        if col == "Tienda":
            continue
        values = pd.to_numeric(pdf_df[col], errors="coerce").fillna(0)
        pdf_df[col] = values.map(lambda x: f"{x:.1f}%" if "%" in col else f"{x:,.0f}")

    data = [list(pdf_df.columns)] + pdf_df.astype(str).values.tolist()
    widths = [70,45,45,43,46,52,48,60,58,55,52,50,54,48]
    table = Table(data, colWidths=widths, repeatRows=1)
    table_style = [
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor(AZUL)),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 6.1),
        ("FONTSIZE", (0,1), (-1,-1), 6.0),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#DDE4F0")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F7F9FC")]),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]

    # Semáforo en porcentaje de ubicación.
    if "% Ubic." in pdf_df.columns:
        pct_col = list(pdf_df.columns).index("% Ubic.")
        for row_idx, pct in enumerate(raw_pct_ubic, start=1):
            if pct < 75:
                color = colors.HexColor("#D71920")
            elif pct >= 90:
                color = colors.HexColor("#008A3B")
            else:
                color = colors.HexColor("#111827")
            table_style.extend([
                ("TEXTCOLOR", (pct_col, row_idx), (pct_col, row_idx), color),
                ("FONTNAME", (pct_col, row_idx), (pct_col, row_idx), "Helvetica-Bold"),
            ])

    table.setStyle(TableStyle(table_style))
    story.append(table)
    story.append(Spacer(1, 7))
    story.append(Paragraph(
        "<b>Ingreso vs Habilitado vs Ubicado por tienda</b>",
        ParagraphStyle("pdf_h3", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#1D1259"), spaceAfter=2),
    ))
    story.append(_pdf_chart(df))

    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buffer.seek(0)
    return buffer.getvalue()

def download_pdf_button(label="Descargar PDF", title="Reporte", subtitle="", kpi_values=None, df=None, key=None):
    if kpi_values is not None and df is not None:
        pdf = build_pdf_report(title, subtitle, kpi_values, df)

        # Extrae la fecha del subtítulo, por ejemplo: "Fecha: 2026-06-28".
        date_match = re.search(r"(\d{4})[-/](\d{2})[-/](\d{2})", str(subtitle))
        if date_match:
            yyyy, mm, dd = date_match.groups()
            date_suffix = f"{dd}-{mm}-{yyyy}"
        else:
            date_suffix = pd.Timestamp.today().strftime("%d-%m-%Y")

        clean_title = re.sub(r"[^A-Za-z0-9ÁÉÍÓÚáéíóúÑñ]+", "_", str(title)).strip("_")
        if clean_title.lower().replace("_", "") in {"pordia", "reporte_pordia"}:
            file_name = f"Reporte_Por_Dia_{date_suffix}.pdf"
        else:
            file_name = f"{clean_title}_{date_suffix}.pdf"

        st.download_button(
            label,
            data=pdf,
            file_name=file_name,
            mime="application/pdf",
            key=key or f"pdf_{clean_title}_{date_suffix}",
        )
    else:
        st.button(label, help="PDF disponible en pestañas con indicadores.")



def build_generic_table_pdf(title, subtitle, df, kpi_values=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=18,
        leftMargin=18,
        topMargin=14,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()
    story = []

    logo_path = ASSETS_DIR / "price_shoes_logo.png"
    logo = RLImage(str(logo_path), width=58, height=34) if logo_path.exists() else Paragraph("<b>Price Shoes</b>", styles["Normal"])
    header_text = Paragraph(
        f"<font name='Helvetica-Bold' color='#1D1259' size='13'>Indicadores Cambios y Muertos</font><br/>"
        f"<font name='Helvetica-Bold' color='#1D1259' size='10'>{title}</font>"
        f"<font name='Helvetica' color='#5B6476' size='8'> | {subtitle}</font>",
        ParagraphStyle("generic_header", parent=styles["Normal"], leading=14),
    )
    header = Table([[logo, header_text]], colWidths=[72, 650], rowHeights=[40])
    header.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(header)

    pink_line = Table([[""]], colWidths=[744], rowHeights=[3])
    pink_line.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor(ROSA)),
    ]))
    story.append(pink_line)
    story.append(Spacer(1, 7))

    if kpi_values:
        cards = [
            _pdf_kpi_card("↻", "Piezas Ingresadas", fmt_num(kpi_values.get("Ingresos", 0)), "Dev + muertos + cajas + probador", ROSA, styles),
            _pdf_kpi_card("✓", "Piezas Acondicionadas", fmt_num(kpi_values.get("Acondicionado", 0)), "Acondicionado", "#5B00D6", styles),
            _pdf_kpi_card("⊕", "Piezas Ubicadas", fmt_num(kpi_values.get("Ubicado", 0)), "Ubicado", "#F59E0B", styles),
            _pdf_kpi_card("⌛", "Pendientes por Ubicar", fmt_num(kpi_values.get("Pendiente", 0)), "Piezas ingresadas - piezas ubicadas", "#05B957", styles),
            _pdf_kpi_card("%", "% Procesado", fmt_pct(kpi_values.get("% Procesado", 0)), "Piezas ubicadas / piezas ingresadas", "#5B00D6", styles),
        ]
        cards_row = Table([cards], colWidths=[148] * 5, rowHeights=[68])
        cards_row.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING", (0,0), (-1,-1), 2),
            ("RIGHTPADDING", (0,0), (-1,-1), 2),
        ]))
        story.append(cards_row)
        story.append(Spacer(1, 7))

    if df is None or df.empty:
        story.append(Paragraph("Sin información para el periodo seleccionado.", styles["Normal"]))
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    out = df.copy()
    numeric_raw = {}
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            numeric_raw[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).tolist()
            values = pd.to_numeric(out[col], errors="coerce").fillna(0)
            if "%" in str(col):
                out[col] = values.map(lambda x: f"{x:.1f}%")
            elif "$" in str(col) or "Importe" in str(col) or "Recuperación" in str(col):
                out[col] = values.map(lambda x: f"${x:,.0f}")
            else:
                out[col] = values.map(lambda x: f"{x:,.0f}")

    max_cols = max(1, len(out.columns))
    widths = [730 / max_cols] * max_cols
    data = [list(out.columns)] + out.astype(str).values.tolist()
    table = Table(data, colWidths=widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor(AZUL)),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 6.2),
        ("FONTSIZE", (0,1), (-1,-1), 5.9),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#DDE4F0")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F7F9FC")]),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]

    # Semáforo solicitado para porcentajes de ubicación:
    # menor de 75% rojo; 90% o más verde.
    for pct_name in ["% Ubic.", "% Ubicado", "% Ubicación"]:
        if pct_name in out.columns and pct_name in numeric_raw:
            col_idx = list(out.columns).index(pct_name)
            for row_idx, pct in enumerate(numeric_raw[pct_name], start=1):
                if pct < 75:
                    color = colors.HexColor("#D71920")
                elif pct >= 90:
                    color = colors.HexColor("#008A3B")
                else:
                    color = colors.HexColor("#111827")
                style_cmds.extend([
                    ("TEXTCOLOR", (col_idx, row_idx), (col_idx, row_idx), color),
                    ("FONTNAME", (col_idx, row_idx), (col_idx, row_idx), "Helvetica-Bold"),
                ])

    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    # Semanal y mensual usan la misma tabla operativa, por lo que se agrega
    # también el gráfico combinado dentro del PDF.
    chart_cols = {"Tienda", "Total", "Habilitadas", "Ubicadas"}
    if chart_cols.issubset(set(df.columns)):
        story.append(Spacer(1, 7))
        story.append(Paragraph(
            "<b>Ingreso vs Habilitado vs Ubicado por tienda</b>",
            ParagraphStyle(
                "generic_chart_title",
                parent=styles["Normal"],
                fontSize=9,
                textColor=colors.HexColor("#1D1259"),
                spaceAfter=2,
            ),
        ))
        story.append(_pdf_chart(df))

    doc.build(story, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buffer.seek(0)
    return buffer.getvalue()


def generic_pdf_button(title, subtitle, df, kpi_values=None, file_name=None, key=None):
    pdf = build_generic_table_pdf(title, subtitle, df, kpi_values)
    if not file_name:
        clean = re.sub(r"[^A-Za-z0-9ÁÉÍÓÚáéíóúÑñ]+", "_", title).strip("_")
        file_name = f"{clean}.pdf"
    st.download_button(
        "Descargar PDF",
        data=pdf,
        file_name=file_name,
        mime="application/pdf",
        key=key or f"pdf_generic_{re.sub(r'[^A-Za-z0-9]', '', title)}",
    )

def login_sidebar():
    if "user" in st.session_state:
        return True

    # Ocultar elementos de Streamlit durante el acceso para que realmente
    # ocupe toda la pantalla útil.
    st.markdown(
        """
        <style>
        header[data-testid="stHeader"],
        [data-testid="stToolbar"],
        [data-testid="stDecoration"],
        #MainMenu,
        footer {
            display: none !important;
        }
        [data-testid="stAppViewContainer"] {
            background:
                linear-gradient(rgba(0,27,28,.78), rgba(0,43,36,.92)),
                radial-gradient(circle at 50% 18%, rgba(255,255,255,.10), transparent 34%),
                linear-gradient(145deg,#061F29,#003D33) !important;
        }
        [data-testid="stMain"] {
            min-height: 100vh !important;
        }
        .block-container {
            max-width: 760px !important;
            min-height: 100vh !important;
            padding: 5vh 1rem 2rem !important;
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        <div class="login-brand-card">
            <div class="login-real-logo">{logo_html()}</div>
            <div class="login-portal-title">Operaciones Ropa</div>
            <div class="login-portal-subtitle">Indicadores</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.form("login_portal_form", clear_on_submit=False):
        nom = st.text_input(
            "Usuario o correo",
            key="login_user",
            placeholder="Ingresa tu usuario o nómina",
        )
        pwd = st.text_input(
            "Contraseña",
            type="password",
            key="login_password",
            placeholder="Ingresa tu contraseña",
        )
        submitted = st.form_submit_button(
            "Iniciar sesión",
            type="primary",
            use_container_width=True,
        )

    if submitted:
        user = get_user(nom, pwd)
        if user:
            st.session_state.user = user
            st.session_state["nav_page"] = "Resumen"
            st.rerun()
        else:
            st.error("Usuario o contraseña incorrectos.")
    return False


def sidebar_data_admin():
    st.sidebar.divider()
    st.sidebar.markdown("## 📁 Fuente de datos")
    meta = {}
    if META_FILE.exists():
        try:
            meta = json.loads(META_FILE.read_text(encoding="utf-8"))
        except Exception:
            meta = {}

    if ACTIVE_FILE.exists():
        st.sidebar.success("Archivo cargado")
        st.sidebar.write(meta.get("nombre_original", ACTIVE_FILE.name))
        st.sidebar.caption(meta.get("fecha_carga", ""))
        if cache_valid():
            cm = json.loads(cache_paths()["meta"].read_text(encoding="utf-8"))
            st.sidebar.caption(f"Procesado: {cm.get('procesado','')}")
        else:
            st.sidebar.warning("Pendiente de procesar")
    else:
        st.sidebar.warning("No hay archivo cargado")

    if st.session_state.get("user", {}).get("permiso") == "Administrador":
        up = st.sidebar.file_uploader("Cargar/Reemplazar Excel", type=["xlsx"])
        if up is not None and st.sidebar.button("Guardar archivo", type="primary"):
            save_uploaded_file(up)
            st.sidebar.success("Archivo guardado. Ahora presiona Procesar archivo activo.")
            st.rerun()

        if ACTIVE_FILE.exists() and not cache_valid():
            if st.sidebar.button("Procesar archivo activo", type="primary", width="stretch"):
                try:
                    process_excel(str(ACTIVE_FILE))
                    st.success("Archivo procesado correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error("No fue posible procesar el archivo.")
                    st.exception(e)
                    st.stop()

        if ACTIVE_FILE.exists() and st.sidebar.button("Borrar archivo persistido"):
            delete_active_file()
            st.rerun()

    st.sidebar.markdown(
        '<div style="background:#EAF1FF;border-radius:12px;padding:16px;margin-top:24px;"><b style="color:#4F46E5;">🛡️ CONFIDENCIAL</b><br>Price Shoes | Operaciones Ropa</div>',
        unsafe_allow_html=True
    )


def render_app_portal():
    user = st.session_state.get("user", {})
    permiso = user.get("permiso", "Consulta")

    render_portal_header()

    st.markdown(
        """
        <div class="portal-section-title">Aplicaciones</div>
        <div class="portal-section-subtitle">
            Selecciona el indicador que deseas consultar.
        </div>
        """,
        unsafe_allow_html=True,
    )

    app_col, admin_col = st.columns([7, 3], vertical_alignment="top")

    with app_col:
        st.markdown(
            """
            <div class="app-tile">
                <div class="app-tile-icon">↻</div>
                <div class="app-tile-copy">
                    <div class="app-tile-title">Cambios y Muertos</div>
                    <div class="app-tile-subtitle">
                        Recuperación, productividad y conversión de mercancía.
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button(
            "Ingresar a Cambios y Muertos",
            key="open_cambios_muertos",
            type="primary",
            use_container_width=True,
        ):
            st.session_state["active_app"] = "Cambios y Muertos"
            st.session_state["nav_page"] = "Resumen"
            st.rerun()

    with admin_col:
        if permiso == "Administrador":
            with st.expander("⚙️ Administración de Cambios y Muertos", expanded=False):
                st.markdown("#### Fuente de datos")
                meta = {}
                if META_FILE.exists():
                    try:
                        meta = json.loads(META_FILE.read_text(encoding="utf-8"))
                    except Exception:
                        meta = {}

                if ACTIVE_FILE.exists():
                    st.success("Archivo cargado")
                    st.caption(meta.get("nombre_original", ACTIVE_FILE.name))
                    st.caption(meta.get("fecha_carga", ""))
                    if cache_valid():
                        st.caption("Estado: procesado")
                    else:
                        st.warning("Estado: pendiente de procesar")
                else:
                    st.warning("No hay archivo cargado")

                up = st.file_uploader(
                    "Cargar o reemplazar Excel",
                    type=["xlsx"],
                    key="portal_upload_excel",
                )

                if up is not None and st.button(
                    "Guardar archivo",
                    key="portal_save_excel",
                    type="primary",
                    use_container_width=True,
                ):
                    save_uploaded_file(up)
                    st.success("Archivo guardado. Ahora procesa el archivo.")
                    st.rerun()

                if ACTIVE_FILE.exists() and not cache_valid():
                    if st.button(
                        "Procesar archivo activo",
                        key="portal_process_excel",
                        type="primary",
                        use_container_width=True,
                    ):
                        try:
                            process_excel(str(ACTIVE_FILE))
                            st.success("Archivo procesado correctamente.")
                            st.rerun()
                        except Exception as exc:
                            st.error("No fue posible procesar el archivo.")
                            st.exception(exc)

                if ACTIVE_FILE.exists() and st.button(
                    "Borrar archivo persistido",
                    key="portal_delete_excel",
                    use_container_width=True,
                ):
                    delete_active_file()
                    st.rerun()
        else:
            st.info("El archivo de datos es administrado por un usuario Administrador.")

# ============================================================
# PÁGINAS
# ============================================================
PAGES = [
    "Resumen", "Por Día", "Reporte Semanal", "Reporte Mensual", "Conversión",
    "Recuperación Económica", "Productividad", "Recorridos", "Ranking", "Macro",
    "Diagnóstico", "Configuración", "Usuarios",
]


def nav_bar():
    current = st.session_state.get("nav_page", PAGES[0])
    if current not in PAGES:
        current = PAGES[0]

    selected = st.radio(
        "Pestañas",
        PAGES,
        index=PAGES.index(current),
        horizontal=True,
        label_visibility="collapsed",
        key="nav_v120_tabs",
    )

    st.session_state["nav_page"] = selected
    return selected


def reliable_data_horizon(op, co):
    """Obtiene el horizonte real sin eliminar la nueva operación de julio.

    Antes se utilizaba únicamente la fecha máxima comercial. Como las hojas
    comerciales cargadas terminaban el 28/06/2026, toda la información de
    `Resultados productividad 2` del 29/06 en adelante quedaba descartada.

    Ahora:
    - se toman fechas válidas tanto de operación como de comercial;
    - se descartan únicamente fechas futuras anómalas;
    - se conserva la información operativa nueva hasta la fecha actual.
    """
    op = normalize_operation_df(op)
    co = normalize_commercial_df(co)

    op_dates = (
        pd.to_datetime(op["Fecha"], errors="coerce").dropna()
        if op is not None and not op.empty and "Fecha" in op.columns
        else pd.Series(dtype="datetime64[ns]")
    )
    co_dates = (
        pd.to_datetime(co["Fecha"], errors="coerce").dropna()
        if co is not None and not co.empty and "Fecha" in co.columns
        else pd.Series(dtype="datetime64[ns]")
    )

    # Tolerancia corta para capturas con diferencia de zona horaria.
    max_allowed = pd.Timestamp.today().normalize() + pd.Timedelta(days=2)

    if not op_dates.empty:
        op_dates = op_dates[op_dates <= max_allowed]
    if not co_dates.empty:
        co_dates = co_dates[co_dates <= max_allowed]

    all_dates = pd.concat(
        [s for s in [op_dates, co_dates] if not s.empty],
        ignore_index=True,
    ) if (not op_dates.empty or not co_dates.empty) else pd.Series(dtype="datetime64[ns]")

    if all_dates.empty:
        today = pd.Timestamp.today().normalize()
        return today, today

    return all_dates.min().normalize(), all_dates.max().normalize()


def reliable_operation(op, co):
    """Conserva la operación válida de ambas hojas y elimina fechas anómalas."""
    op = normalize_operation_df(op)
    if op is None or op.empty:
        return op

    min_date, max_date = reliable_data_horizon(op, co)
    dates = pd.to_datetime(op["Fecha"], errors="coerce")

    out = op[
        dates.notna()
        & (dates >= min_date)
        & (dates <= max_date)
    ].copy()

    if "Fecha" in out.columns and not out.empty:
        out["Fecha"] = pd.to_datetime(out["Fecha"], errors="coerce").dt.normalize()
        out["Semana ISO"] = out["Fecha"].dt.isocalendar().week.astype(int)
        out["Año ISO"] = out["Fecha"].dt.isocalendar().year.astype(int)
        out["Mes"] = out["Fecha"].dt.to_period("M").astype(str)

    return out


def available_iso_weeks(op, co):
    """Devuelve pares (año ISO, semana ISO) válidos y ordenados."""
    op = reliable_operation(op, co)
    if op is None or op.empty or "Fecha" not in op.columns:
        return []

    fechas = pd.to_datetime(op["Fecha"], errors="coerce")
    valid = fechas.notna()
    if not valid.any():
        return []

    iso = fechas[valid].dt.isocalendar()
    pairs = (
        pd.DataFrame({
            "iso_year": iso["year"].astype(int).to_numpy(),
            "iso_week": iso["week"].astype(int).to_numpy(),
        })
        .drop_duplicates()
        .sort_values(["iso_year", "iso_week"])
    )

    # name=None evita que pandas cambie nombres de columnas con guion bajo
    # al convertirlas en namedtuples.
    return [(int(year), int(week)) for year, week in pairs.itertuples(index=False, name=None)]


def last_four_iso_week_ranges(op, co=None):
    """Cuatro semanas ISO consecutivas terminando en la última fecha real cargada."""
    op = reliable_operation(op, co)
    if op is None or op.empty:
        return []

    _, latest = reliable_data_horizon(op, co)
    current_monday = latest - pd.Timedelta(days=int(latest.weekday()))
    ranges = []
    for offset in [3, 2, 1, 0]:
        monday = current_monday - pd.Timedelta(weeks=offset)
        sunday = monday + pd.Timedelta(days=6)
        iso = monday.isocalendar()
        ranges.append({
            "iso_year": int(iso.year),
            "iso_week": int(iso.week),
            "start": monday,
            "end": sunday,
        })
    return ranges


def executive_week_cards(op, co):
    op = reliable_operation(op, co)
    co = normalize_commercial_df(co)
    week_ranges = last_four_iso_week_ranges(op, co)
    if not week_ranges:
        return

    html = '<div style="margin:18px 0 8px 0;font-size:24px;font-weight:900;color:#3E4095;">📊 Resumen Ejecutivo</div>'
    html += '<div class="week-card-grid">'
    prev_ing = None
    prev_hab = None
    prev_ub = None

    for wr in week_ranges:
        df = table_by_store(op, co, wr["start"], wr["end"], PROJECT_STORES)
        ingresos = float(pd.to_numeric(df["Total"], errors="coerce").fillna(0).sum())
        hab = float(pd.to_numeric(df["Habilitadas"], errors="coerce").fillna(0).sum())
        ub = float(pd.to_numeric(df["Ubicadas"], errors="coerce").fillna(0).sum())

        week_mask = (
            (pd.to_datetime(op["Fecha"], errors="coerce") >= wr["start"])
            & (pd.to_datetime(op["Fecha"], errors="coerce") <= wr["end"])
        )
        if "Actividad" in op.columns:
            actividad = op["Actividad"].map(norm_text)
            recorridos = int(
                (
                    week_mask
                    & actividad.str.contains(r"\bRECORRIDO(S)?\b", regex=True, na=False)
                ).sum()
            )
        else:
            recorridos = 0

        def delta(cur, prev):
            if prev is None or prev == 0:
                return "—", "#6B7280"
            d = (cur - prev) / prev * 100
            icon = "▲" if d >= 0 else "▼"
            color = "#00A651" if d >= 0 else "#EC004F"
            return f"{icon} {abs(d):.1f}%", color

        d_ing, c_ing = delta(ingresos, prev_ing)
        d_hab, c_hab = delta(hab, prev_hab)
        d_ub, c_ub = delta(ub, prev_ub)

        html += (
            f'<div class="week-card">'
            f'<div class="week-card-head">Sem {wr["iso_week"]}</div>'
            f'<div class="week-row"><span>INGRESOS</span><b>{ingresos:,.0f}</b><em style="color:{c_ing};">{d_ing}</em></div>'
            f'<div class="week-row"><span>ACONDICIONADO</span><b>{hab:,.0f}</b><em style="color:{c_hab};">{d_hab}</em></div>'
            f'<div class="week-row"><span>UBICADO</span><b>{ub:,.0f}</b><em style="color:{c_ub};">{d_ub}</em></div>'
            f'<div class="week-row"><span>RECORRIDOS</span><b>{recorridos:,.0f}</b><em>—</em></div>'
            f'</div>'
        )
        prev_ing, prev_hab, prev_ub = ingresos, hab, ub

    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


def page_resumen(op, co):
    op = reliable_operation(op, co)
    co = normalize_commercial_df(co)
    st.markdown("## Dashboard Ejecutivo")
    st.caption("Vista general acumulada de indicadores principales.")

    if op is None or op.empty:
        st.info("Sin información operativa.")
        return

    op_dates = pd.to_datetime(op["Fecha"], errors="coerce").dropna()
    co_dates = (
        pd.to_datetime(co["Fecha"], errors="coerce").dropna()
        if co is not None and not co.empty and "Fecha" in co.columns
        else pd.Series(dtype="datetime64[ns]")
    )

    all_dates = pd.concat(
        [s for s in [op_dates, co_dates] if not s.empty],
        ignore_index=True,
    )
    if all_dates.empty:
        st.info("Sin fechas válidas.")
        return

    # Las tarjetas superiores son acumuladas desde el primer registro hasta el último.
    start = all_dates.min().normalize()
    end = all_dates.max().normalize()
    df = table_by_store(op, co, start, end, PROJECT_STORES)

    kpis(summary_from_table(df))
    st.caption(
        f"Acumulado del {start.strftime('%d/%m/%Y')} al {end.strftime('%d/%m/%Y')}."
    )

    # Las tarjetas inferiores conservan el análisis de las últimas cuatro semanas.
    executive_week_cards(op, co)
    combined_chart(df, "Ingreso vs Habilitado vs Ubicado por tienda — Acumulado")


def page_por_dia(op, co):
    op = reliable_operation(op, co)
    co = normalize_commercial_df(co)
    st.markdown("## Por Día")
    st.caption("Ingresos, pendientes y avance por tienda.")

    if co is not None and not co.empty and "Fecha" in co.columns:
        default_date = pd.to_datetime(co["Fecha"].max()).date()
    elif op is not None and not op.empty:
        default_date = pd.to_datetime(op["Fecha"].max()).date()
    else:
        default_date = date.today()

    d = st.date_input("Fecha", value=default_date, key="dia_fecha")
    d_ts = parse_date(d)
    df = table_by_store(op, co, d_ts, d_ts, PROJECT_STORES)

    op_count = len(op[pd.to_datetime(op["Fecha"], errors="coerce").dt.normalize().eq(d_ts)]) if op is not None and not op.empty else 0
    co_dia = filter_commercial_by_date(co, d_ts, d_ts, PROJECT_STORES)
    dev_sum = co_dia["Dev_Pzs"].sum() if co_dia is not None and not co_dia.empty and "Dev_Pzs" in co_dia.columns else 0
    st.caption(f"Registros detectados: operación {op_count:,} | Dev Pzs mensual {dev_sum:,.0f}")

    resumen = summary_from_table(df)
    kpis(resumen)
    download_pdf_button("Descargar PDF", "Por Dia", f"Fecha: {pd.to_datetime(d_ts).strftime('%Y-%m-%d')}", resumen, df, key="pdf_por_dia")
    panel("Tabla por tienda - Por Día", df, height=360)
    combined_chart(df, "Ingreso vs Habilitado vs Ubicado por tienda", income_column="Ingresos periodo")


def page_semanal(op, co):
    op = reliable_operation(op, co)
    st.markdown("## Reporte Semanal")
    tiendas = st.multiselect("Tiendas", PROJECT_STORES, default=PROJECT_STORES, key="sem_tiendas")

    week_pairs = available_iso_weeks(op, co)
    if not week_pairs:
        st.info("Sin semanas válidas detectadas.")
        return

    labels = [f"{year}-Sem {week:02d}" for year, week in week_pairs]
    selected_label = st.selectbox("Semana ISO", labels, index=len(labels)-1, key="sem_iso")
    year, week = week_pairs[labels.index(selected_label)]

    dates = pd.to_datetime(op["Fecha"], errors="coerce")
    iso = dates.dt.isocalendar()
    mask = (iso.year.astype(int) == year) & (iso.week.astype(int) == week)
    week_dates = dates[mask]
    if week_dates.empty:
        st.info("Sin fechas para la semana seleccionada.")
        return

    start, end = week_dates.min().normalize(), week_dates.max().normalize()
    df = table_by_store(op, co, start, end, tiendas, carryover_mode="previous_sunday")
    resumen = summary_from_table(df, income_column="Total")
    kpis(resumen)
    st.caption("La base semanal incluye el saldo pendiente acumulado al cierre del domingo anterior.")

    generic_pdf_button(
        f"Reporte Semanal - Semana {week}",
        f"Periodo: {start.strftime('%d-%m-%Y')} al {end.strftime('%d-%m-%Y')}",
        df,
        resumen,
        file_name=f"Reporte_Semanal_Semana_{week:02d}_{year}.pdf",
        key=f"pdf_sem_{year}_{week}",
    )
    panel(f"Tabla por tienda - Semana {week}", df, height=360)
    combined_chart(df, f"Ingreso vs Habilitado vs Ubicado - Semana {week}")


def page_mensual(op, co):
    op = reliable_operation(op, co)
    st.markdown("## Reporte Mensual")
    tiendas = st.multiselect("Tiendas", PROJECT_STORES, default=PROJECT_STORES, key="mes_tiendas")
    meses = sorted(op["Mes"].dropna().unique().tolist()) if op is not None and not op.empty else []
    if not meses:
        st.info("Sin meses detectados.")
        return
    m = st.selectbox("Mes", meses, index=len(meses)-1, key="mes_select")
    dates = pd.to_datetime(op.loc[op["Mes"].eq(m), "Fecha"], errors="coerce").dropna()
    if dates.empty:
        st.info("Sin fechas para el mes seleccionado.")
        return
    start, end = dates.min().normalize(), dates.max().normalize()
    df = table_by_store(op, co, start, end, tiendas, carryover_mode="none")
    resumen = summary_from_table(df, income_column="Ingresos periodo")
    kpis(resumen)
    st.caption("Los ingresos mensuales consideran únicamente movimientos del mes; el pendiente se suma por tienda.")
    generic_pdf_button(
        f"Reporte Mensual - {m}",
        f"Periodo: {start.strftime('%d-%m-%Y')} al {end.strftime('%d-%m-%Y')}",
        df, resumen,
        file_name=f"Reporte_Mensual_{m}.pdf",
        key=f"pdf_mes_{m}",
    )
    panel(f"Tabla por tienda - Mes {m}", df, height=360)
    combined_chart(df, f"Ingreso vs Habilitado vs Ubicado - Mes {m}", income_column="Ingresos periodo")


def page_conversion(op, co):
    st.markdown("## Conversión Semanal Dev → Venta")
    st.caption("Se respeta semana ISO por tienda. La conversión real requiere ventas por ID/color/semana en el archivo.")
    if co.empty:
        st.info("Sin información comercial mensual.")
        return
    weeks = sorted(co["Semana ISO"].dropna().unique().tolist())
    w = st.multiselect("Semana ISO", weeks, default=weeks[-1:])
    df = co[co["Semana ISO"].isin(w)].groupby(["Semana ISO","Tienda"], as_index=False).agg({"Dev_Pzs":"sum","Vta_Pzs":"sum","Vta_Imp":"sum"})
    df["Pend. Conv."] = df["Dev_Pzs"] - df["Vta_Pzs"]
    df["% Conv."] = np.where(df["Dev_Pzs"]>0, df["Vta_Pzs"]/df["Dev_Pzs"]*100, 0)
    df = df.rename(columns={"Dev_Pzs":"Dev Pzs", "Vta_Pzs":"Conv. Pzs", "Vta_Imp":"Conv. $"})
    generic_pdf_button("Conversión Semanal Dev a Venta", f"Semanas: {w}", df, file_name="Reporte_Conversion.pdf", key="pdf_conversion")
    panel("Detalle Conversión", df, height=420)


def page_recuperacion(op, co):
    st.markdown("## Recuperación Económica")
    if co.empty:
        st.info("Sin información comercial mensual.")
        return
    weeks = sorted(co["Semana ISO"].dropna().unique().tolist())
    w = st.multiselect("Semana ISO", weeks, default=weeks[-1:], key="rec_sem")
    df = co[co["Semana ISO"].isin(w)].groupby(["Semana ISO","Tienda"], as_index=False).agg({"Vta_Imp":"sum","Dev_Pzs":"sum"})
    df = df.rename(columns={"Vta_Imp":"Recuperación $", "Dev_Pzs":"Dev Pzs"})
    generic_pdf_button("Recuperación Económica", f"Semanas: {w}", df, file_name="Reporte_Recuperacion_Economica.pdf", key="pdf_recuperacion")
    panel("Recuperación Económica", df, height=420)


def page_productividad(op, co):
    st.markdown("## Productividad")
    if op.empty:
        st.info("Sin operación.")
        return
    c1, c2 = st.columns(2)
    with c1:
        start = st.date_input("Fecha inicio", value=pd.to_datetime(op["Fecha"].min()).date(), key="prod_ini")
    with c2:
        end = st.date_input("Fecha final", value=pd.to_datetime(op["Fecha"].max()).date(), key="prod_fin")
    tiendas = st.multiselect("Tienda", PROJECT_STORES, default=PROJECT_STORES, key="prod_tiendas")
    o = split_operation(op)
    o = o[(o["Fecha"] >= pd.to_datetime(start)) & (o["Fecha"] <= pd.to_datetime(end))]
    o = filter_stores(o, tiendas)
    name_col = "Nombre Real" if "Nombre Real" in o.columns else "Nombre"
    df = o.groupby(["Tienda", name_col], as_index=False).agg({"Piezas":"sum", "Habilitadas":"sum", "Ubicadas":"sum", "Recolectadas":"sum"})
    df = df.rename(columns={name_col:"Colaborador"})
    df_pdf = df.sort_values("Piezas", ascending=False)
    generic_pdf_button("Productividad por colaborador", f"Periodo: {start} al {end}", df_pdf, file_name="Reporte_Productividad.pdf", key="pdf_productividad")
    panel("Productividad por colaborador", df_pdf, height=500)


def page_recorridos(op, co):
    st.markdown("## Recorridos")
    if op.empty:
        return
    o = op[op["Actividad"].map(norm_text).str.contains("RECORRIDO|RECOLECCION|RECOLECCIÓN", na=False)]
    df = o.groupby(["Semana ISO","Tienda"], as_index=False).size().rename(columns={"size":"Recorridos"})
    df["Meta"] = 47
    df["% Cumplimiento"] = df["Recorridos"] / 47 * 100
    generic_pdf_button("Recorridos", "Detalle por semana y tienda", df, file_name="Reporte_Recorridos.pdf", key="pdf_recorridos")
    panel("Recorridos por semana", df, height=420)


def page_ranking(op, co):
    st.markdown("## Ranking")
    if op.empty:
        return
    o = split_operation(op)
    df = o.groupby("Tienda", as_index=False).agg({"Piezas":"sum", "Habilitadas":"sum", "Ubicadas":"sum"})
    df["Score"] = (df["Habilitadas"] + df["Ubicadas"]) / df["Piezas"].replace(0, np.nan) * 100
    df["Score"] = df["Score"].fillna(0)
    df_rank = df.sort_values("Score", ascending=False)
    generic_pdf_button("Ranking de tiendas", "Clasificación por score", df_rank, file_name="Reporte_Ranking.pdf", key="pdf_ranking")
    panel("Ranking de tiendas", df_rank, height=420)


def page_macro(op, co):
    st.markdown("## Macro")
    page_resumen(op, co)


def page_diagnostico(op, co, diag):
    st.markdown("## Diagnóstico")
    st.info("Homologación v10.13: acepta encabezado Tienda/Tiendas; Guadalajara Miravalle/Miravalle => Miravalle; Guadalajara/Guadalajara Atemajac/Atemajac => Atemajac.")
    st.info("Homologación v10.12: Guadalajara Miravalle/Miravalle => Miravalle; Guadalajara/Guadalajara Atemajac/Atemajac => Atemajac.")
    st.write(f"Operación: {len(op):,} registros")
    st.write(f"Comercial mensual Dev Pzs: {len(co):,} registros agrupados | Dev Pzs total: {co['Dev_Pzs'].sum() if not co.empty else 0:,.0f}")
    if diag is not None and not diag.empty and "Tipo" in diag.columns:
        diag_op = diag[diag["Tipo"].astype(str).str.contains("Histórica|Nueva|Consolidado", case=False, na=False)]
        if not diag_op.empty:
            panel("Diagnóstico operativo — unión de hojas", diag_op, height=260)
    panel("Diagnóstico de hojas", diag, height=420)
    if not co.empty:
        _co_diag = normalize_commercial_df(co)
        dev_diag = (
            _co_diag.groupby(["Fecha", "Tienda"], as_index=False)[["Dev_Pzs", "Vta_Pzs", "Vta_Imp"]]
            .sum()
            .sort_values(["Fecha", "Tienda"])
        )
        dev_diag["Fecha"] = pd.to_datetime(dev_diag["Fecha"], errors="coerce").dt.strftime("%Y-%m-%d")
        panel("Validación Comercial por fecha y tienda", dev_diag.tail(300), height=520)
        ecatepec_2806 = dev_diag[(dev_diag["Tienda"].eq("Ecatepec")) & (dev_diag["Fecha"].eq(pd.Timestamp("2026-06-28")))]
        if not ecatepec_2806.empty:
            st.success(f"Validación Ecatepec 28/06/2026 Dev Pzs: {ecatepec_2806['Dev_Pzs'].sum():,.0f}")



def page_configuracion():
    st.markdown("## Configuración")
    st.info("Configuración de metas y orden de pestañas en preparación modular.")
    st.write("Meta productividad diaria: 784")
    st.write("Meta recorridos semanal: 47")


def page_usuarios():
    st.markdown("## Usuarios")
    if st.session_state.get("user", {}).get("permiso") != "Administrador":
        st.warning("Sólo administrador.")
        return

    with st.form("crear_usuario"):
        st.subheader("Crear / actualizar usuario")
        nom = st.text_input("Nómina / Usuario")
        nombre = st.text_input("Nombre")
        permiso = st.selectbox("Tipo de permiso", ["Consulta", "Administrador"])
        pwd = st.text_input("Contraseña", type="password")
        submitted = st.form_submit_button("Guardar usuario", type="primary")
        if submitted and nom and nombre and pwd:
            upsert_user(nom, nombre, permiso, pwd)
            st.success("Usuario guardado.")
            st.rerun()

    users = list_users()
    panel("Usuarios registrados", users, height=360)

    del_nom = st.text_input("Nómina a eliminar")
    if st.button("Eliminar usuario") and del_nom:
        delete_user(del_nom)
        st.success("Usuario eliminado.")
        st.rerun()


# ============================================================
# MAIN
# ============================================================
apply_styles()

if not login_sidebar():
    st.stop()

if "active_app" not in st.session_state:
    st.session_state["active_app"] = None

if not st.session_state.get("active_app"):
    render_app_portal()
    st.stop()

render_header()
page = nav_bar()

if not ACTIVE_FILE.exists():
    st.warning("El Administrador debe cargar el archivo desde el menú principal de aplicaciones.")
    st.stop()

if not cache_valid():
    st.warning("El archivo está cargado, pero aún no está procesado. Regresa al menú principal para procesarlo.")
    st.stop()

op_all, co_all, diag_df = read_cache(ACTIVE_FILE.stat().st_mtime)

ROUTES = {
    "Resumen": lambda: page_resumen(op_all, co_all),
    "Por Día": lambda: page_por_dia(op_all, co_all),
    "Reporte Semanal": lambda: page_semanal(op_all, co_all),
    "Reporte Mensual": lambda: page_mensual(op_all, co_all),
    "Conversión": lambda: page_conversion(op_all, co_all),
    "Recuperación Económica": lambda: page_recuperacion(op_all, co_all),
    "Productividad": lambda: page_productividad(op_all, co_all),
    "Recorridos": lambda: page_recorridos(op_all, co_all),
    "Ranking": lambda: page_ranking(op_all, co_all),
    "Macro": lambda: page_macro(op_all, co_all),
    "Diagnóstico": lambda: page_diagnostico(op_all, co_all, diag_df),
    "Configuración": page_configuracion,
    "Usuarios": page_usuarios,
}

ROUTES.get(page, lambda: page_resumen(op_all, co_all))()

st.markdown(
    '<div class="footer">CONFIDENCIAL | Price Shoes | Operaciones Ropa</div>',
    unsafe_allow_html=True,
)
